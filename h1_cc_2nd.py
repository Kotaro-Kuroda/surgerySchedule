import copy
import os
import time
import graph
import pulp
import numpy as np
from scipy.stats import norm
import instance
import openpyxl as px
import random

home = os.environ['HOME']
directory_path = home + '/Documents/data-july'
path = directory_path + '/operations_with_jisseki_remake.csv'
surgeon_info_path = directory_path + '/surgeon_info.csv'
surgeries_path = directory_path + '/surgeries.csv'
distribution_path = home + '/Documents/surgerySchedule/distribution.csv'
file_path = home + '/Documents/surgerySchedule/room_schedule_h1_cc.xlsx'
file_path2 = home + '/Documents/surgerySchedule/surgeon_schedule_h1_cc.xlsx'
save_path = '/Users/kurodakotaro/Documents/image/CCP/'
num_surgery = 10
num_date = 1
seed = 1

alpha = 0.95
m1 = 1
m2 = 3
m3 = 2
m4 = 4
m5 = 30
m6 = 2
m7 = 0.5
m8 = 1
m9 = 10
m10 = 1
initial = time.time()

T = 480
PT = 15
M = 99999
ope_time = 9
O_max = 120

waiting_list = []

solver = pulp.CPLEX_CMD(msg=False)


def get_not_available_room_dict(lst_surgery, lst_room, dict_surgery_id_room):
    dict_not_available_room = {}
    for surgery in lst_surgery:
        surgery_type_list = surgery.get_surgery_type()
        available_room = set(lst_room)
        for surgery_type in surgery_type_list:
            if surgery_type in dict_surgery_id_room.keys():
                available_room = available_room & set(dict_surgery_id_room[surgery_type])
        not_available_room = set(lst_room) - available_room
        dict_not_available_room[surgery] = list(not_available_room)
    return dict_not_available_room


def get_sigma(mean, variance):
    return np.log(variance / (mean ** 2) + 1)


def get_mu(mean, variance):
    return np.log(mean) - get_sigma(mean, variance) / 2


def get_right_hand(prob, mean, variance):
    sigma = get_sigma(mean, variance)
    mu = get_mu(mean, variance)
    return np.exp(np.sqrt(sigma) * norm.ppf(prob) + mu)


def calc_sigma(x_vec, mean_vec, variance_vec, lst_surgery):
    return np.log(sum(variance_vec[s] * x_vec[s] for s in lst_surgery) + (
        sum(mean_vec[s] * x_vec[s] for s in lst_surgery)) ** 2) - 2 * np.log(
        sum(mean_vec[s] * x_vec[s] for s in lst_surgery))


def grad_sigma(x_vec, mean_vec, variance_vec, lst_surgery):
    grad = {}
    sum_mean = sum(mean_vec[s] * x_vec[s] for s in lst_surgery)
    sum_variance = sum(variance_vec[s] * x_vec[s] for s in lst_surgery)
    for s in lst_surgery:
        grad[s] = (variance_vec[s] + 2 * sum_mean * mean_vec[s]) / (sum_variance + sum_mean ** 2) - 2 * (
            mean_vec[s]) / sum_mean
    return grad


def log_normal_cc(x_vec, mean_vec, variance_vec, lst_surgery, ot_val):
    if set(x_vec.values()) == {0}:
        return -np.inf
    else:
        return np.log(sum(mean_vec[s] * x_vec[s] for s in lst_surgery)) - 1 / 2 * calc_sigma(x_vec, mean_vec,
                                                                                             variance_vec,
                                                                                             lst_surgery) + norm.ppf(
            alpha) * np.sqrt(calc_sigma(x_vec, mean_vec, variance_vec, lst_surgery)) - np.log(ot_val + T)


def grad_log_normal_cc(x_vec, mean_vec, variance_vec, lst_surgery, ot_val):
    grad_dict = {}
    sum_mean = sum(mean_vec[s1] * x_vec[s1] for s1 in lst_surgery)
    grad_s = grad_sigma(x_vec, mean_vec, variance_vec, lst_surgery)
    sig = calc_sigma(x_vec, mean_vec, variance_vec, lst_surgery)
    if set(x_vec.values()) == {0}:
        for s in lst_surgery:
            grad_dict[s] = -np.inf
    else:
        for s in lst_surgery:
            grad_dict[s] = mean_vec[s] / sum_mean - 1 / 2 * grad_s[s] + norm.ppf(alpha) * 1 / 2 * 1 / np.sqrt(sig) * \
                           grad_s[s]
    grad_dict['ot'] = - 1 / (ot_val + T)
    return grad_dict


def long_surgery(surgery, distribution_dict):
    total_mu = distribution_dict[surgery.get_group(), 'total_mu']
    total_sigma = distribution_dict[surgery.get_group(), 'total_sigma']
    return np.exp(total_mu + norm.ppf(alpha) * np.sqrt(total_sigma)) > T + O_max


def max_surgery_time(lst_surgery, distribution_dict):
    max_time = 0
    for surgery in lst_surgery:
        total_mu = distribution_dict[surgery.get_group(), 'total_mu']
        total_sigma = distribution_dict[surgery.get_group(), 'total_sigma']
        total_time = np.exp(total_mu + norm.ppf(alpha) * np.sqrt(total_sigma))
        if total_time > max_time:
            max_time = total_time
    return max_time


def initialize_parameter(list_surgery, list_room, list_surgeon, list_date):
    x_val = {}
    q_val = {}
    ot_val = {}
    wt_val = {}
    ts_val = {}
    for surgery in list_surgery:
        for d in list_date:
            for r in list_room:
                x_val[surgery, r, d] = 0
            for surgeon in list_surgeon:
                q_val[surgery, surgeon, d] = 0
        ts_val[surgery] = 0
    for r in list_room:
        for d in list_date:
            ot_val[r, d] = 0
    for surgeon in list_surgeon:
        for d in list_date:
            wt_val[surgeon, d] = 0
    return x_val, q_val, ot_val, wt_val, ts_val


def sort_surgery(list_surgery):
    for i in range(len(list_surgery) - 1):
        for j in range(i + 1, len(list_surgery)):
            surgery1 = list_surgery[i]
            surgery2 = list_surgery[j]
            if surgery1.get_priority() < surgery2.get_priority():
                temp = surgery1
                list_surgery[i] = surgery2
                list_surgery[j] = temp
    return list_surgery


def get_empty_room(room, date, list_surgery, x_val):
    return not sum(x_val[surgery, room, date] for surgery in list_surgery) > 0


def get_scheduled_surgery_info(surgery, list_room, list_surgeon, list_date, x_val, q_val):
    scheduled_date = 0
    scheduled_surgeon = 0
    scheduled_room = 0
    for d in list_date:
        for surgeon in list_surgeon:
            if q_val[surgery, surgeon, d] == 1:
                scheduled_surgeon = surgeon
                scheduled_date = d
        for r in list_room:
            if x_val[surgery, r, d] == 1:
                scheduled_room = r
    return scheduled_room, scheduled_surgeon.get_surgeon_id(), scheduled_date


def main():
    surgery_instance = instance.SurgeryInstance(path, surgeon_info_path, num_surgery, num_date, seed, surgeries_path,
                                                distribution_path)
    list_room, list_surgery, list_surgeon, list_date, list_group = surgery_instance.get_sets()
    dict_surgery_id_room = surgery_instance.get_room_surgery_dict()
    distribution_dict = surgery_instance.get_dict_distribution()
    dict_not_available_room = get_not_available_room_dict(list_surgery, list_room, dict_surgery_id_room)

    x_val, q_val, ot_val, wt_val, ts_val = initialize_parameter(list_surgery, list_room, list_surgeon, list_date)
    list_surgery_copied = copy.copy(list_surgery)
    L = []
    x = {}
    q = {}
    y = {}
    z = {}
    ts = {}
    msR = {}
    ot = {}
    tsS = {}
    msS = {}
    wt = {}
    ost = {}
    ns = {}
    ot2 = {}
    ar = {}
    scheduled_surgery = []
    start = time.time()
    for d in list_date:
        print("day={:}".format(d))
        feasibility_criteria = False
        if len(list_surgery_copied) == 0:
            break
        for surgery in list_surgery_copied[:]:
            if surgery.get_release_date() <= d:
                if surgery not in waiting_list:
                    waiting_list.append(surgery)
                    list_surgery_copied.remove(surgery)

        l = 0
        while not feasibility_criteria:
            l += 1
            print("loop={:}".format(l))
            for surgery in waiting_list:
                for r in list_room:
                    x[surgery, r, d, l] = pulp.LpVariable(
                        "x({:},{:},{:},{:})".format(surgery.get_surgery_id(), r, d, l), cat='Binary')

            for surgery in waiting_list:
                for surgeon in list_surgeon:
                    q[surgery, surgeon, d, l] = pulp.LpVariable(
                        "q({:},{:},{:},{:})".format(surgery.get_surgery_id(), surgeon.get_surgeon_id(), d, l),
                        cat='Binary')

            for r in list_room:
                ar[r, d, l] = pulp.LpVariable('ar[{:}, {:}, {:}]'.format(r, d, l), lowBound=0, cat='Continuous')

            for r in list_room:
                ot[r, d, l] = pulp.LpVariable("ot({:},{:},{:})".format(r, d, l), lowBound=0, cat='Continuous')

            for surgeon in list_surgeon:
                ost[surgeon, d, l] = pulp.LpVariable("ost({:},{:},{:})".format(surgeon.get_surgeon_id(), d, l),
                                                     lowBound=0, cat='Continuous')

            ns[d, l] = pulp.LpVariable("ns({:},{:})".format(d, l), lowBound=0, cat='Continuous')

            operating_room_planning = pulp.LpProblem("ORP" + str(l), pulp.LpMinimize)
            objective = pulp.lpSum(
                (m1 * (d - surgery.get_release_date()) + m2 * max(d - surgery.get_due_date(), 0)) * x[
                    surgery, r, d, l] * surgery.get_priority() for surgery in waiting_list for r in list_room)
            objective += pulp.lpSum((m3 * (surgery.get_due_date() - surgery.get_release_date()) + m4 * (
                    len(list_date) + 1 - surgery.get_due_date())) * surgery.get_priority() * (
                                            1 - pulp.lpSum(x[surgery, r, d, l] for r in list_room)) for surgery in
                                    waiting_list)
            objective += m5 * pulp.lpSum(
                surgery.get_priority() * (1 - pulp.lpSum(x[surgery, r, d, l] for r in list_room)) for surgery in
                waiting_list)
            objective += m6 * pulp.lpSum(ot[r, d, l] for r in list_room)
            operating_room_planning += objective
            list_long_surgery = [surgery for surgery in waiting_list if long_surgery(surgery, distribution_dict)]
            for surgery in list_long_surgery:
                print(surgery.get_surgery_id())
            list_long_surgery = sort_surgery(list_long_surgery)
            operating_room_planning += pulp.lpSum(x[list_long_surgery[0], r, d, l] for r in list_room) == 1
            for surgery in waiting_list:
                operating_room_planning += pulp.lpSum(x[surgery, r, d, l] for r in list_room) <= 1
            for surgery in waiting_list:
                operating_room_planning += pulp.lpSum(q[surgery, surgeon, d, l] for surgeon in list_surgeon) <= 1

            for surgery in waiting_list:
                operating_room_planning += pulp.lpSum(x[surgery, r, d, l] for r in list_room) == pulp.lpSum(
                    q[surgery, surgeon, d, l] for surgeon in list_surgeon)
            """
            for surgeon in list_surgeon:
                operating_room_planning += ost[surgeon, d, l] >= pulp.lpSum(
                    q[surgery, surgeon, d, l] for surgery in waiting_list) - len(waiting_list) / len(list_surgeon)
                operating_room_planning += - ost[surgeon, d, l] <= pulp.lpSum(
                    q[surgery, surgeon, d, l] for surgery in waiting_list) - len(waiting_list) / len(list_surgeon)
            """
            for r in list_room:
                a = {}
                ot1 = 0
                if l == 1:
                    ot1 = random.randint(0, O_max)
                    # index = random.randint(0, len(waiting_list) - 1)
                    for s in waiting_list:
                        a[s] = np.random.randint(0, 1)
                else:
                    for surgery in waiting_list:
                        a[surgery] = x_val[surgery, r, d, l - 1]
                        ot1 = ot_val[r, d, l - 1]
                mean_vec = {}
                variance_vec = {}
                for surgery in waiting_list:
                    mean_vec[surgery] = distribution_dict[surgery.get_group(), 'total_mean']
                    variance_vec[surgery] = distribution_dict[surgery.get_group(), 'total_variance']
                if set(a.values()) == {0}:
                    min_mean = min(list(mean_vec.values()))
                    index = [k for k, v in mean_vec.items() if v == min_mean][0]
                    a[index] = 1
                grad_dict = grad_log_normal_cc(a, mean_vec, variance_vec, waiting_list, ot1)
                approx = log_normal_cc(a, mean_vec, variance_vec, waiting_list, ot1) + pulp.lpSum(
                    grad_dict[surgery] * (x[surgery, r, d, l] - a[surgery]) for surgery in waiting_list) + grad_dict[
                             'ot'] * (ot[r, d, l] - ot1)
                operating_room_planning += approx <= 0
                operating_room_planning += ot[r, d, l] <= O_max + M * x[list_long_surgery[0], r, d, l]
            """
            for r in list_room:
                operating_room_planning += ar[r, d, l] >= pulp.lpSum(
                    x[surgery, r, d, l] for surgery in waiting_list) - len(waiting_list) / len(list_room)
                operating_room_planning += -ar[r, d, l] <= pulp.lpSum(
                    x[surgery, r, d, l] for surgery in waiting_list) - len(waiting_list) / len(list_room)
            """
            """
            operating_room_planning += ns[d, l] >= pulp.lpSum(
                x[surgery, r, d, l] for surgery in waiting_list for r in list_room) - len(list_surgery) / len(list_date)
            operating_room_planning += -ns[d, l] <= pulp.lpSum(
                x[surgery, r, d, l] for surgery in waiting_list for r in list_room) - len(list_surgery) / len(list_date)
            """
            for surgery in waiting_list:
                for r in dict_not_available_room[surgery]:
                    operating_room_planning += x[surgery, r, d, l] == 0

            for surgery in waiting_list:
                group = surgery.get_group()
                for surgeon in list_surgeon:
                    if float(surgeon.get_dict_group()[group]) <= 0:
                        operating_room_planning += q[surgery, surgeon, d, l] == 0

            if l > 1:
                for i in range(1, l):
                    for surgery1 in waiting_list:
                        for surgery2 in waiting_list:
                            if surgery1 != surgery2:
                                operating_room_planning += pulp.lpSum(x[surgery1, r, d, l] for r in list_room if
                                                                      x_val[surgery1, r, d, i] == 1) + pulp.lpSum(
                                    x[surgery2, r, d, l] for r in list_room if
                                    x_val[surgery2, r, d, i] == 1) + pulp.lpSum(
                                    q[surgery1, surgeon, d, l] for surgeon in list_surgeon if
                                    q_val[surgery1, surgeon, d, i] == 1) + pulp.lpSum(
                                    q[surgery2, surgeon, d, l] for surgeon in list_surgeon if
                                    q_val[surgery2, surgeon, d, i] == 1) <= 3

            result_status = operating_room_planning.solve(solver)
            print('ORP={:}'.format(pulp.LpStatus[result_status]))
            for surgery in waiting_list:
                for r in list_room:
                    x_val[surgery, r, d, l] = round(x[surgery, r, d, l].value())
                    """
                    if x_val[surgery, r, d, l] == 1:
                        print("r={:}".format(r))
                    """
                for surgeon in list_surgeon:
                    q_val[surgery, surgeon, d, l] = round(q[surgery, surgeon, d, l].value())
                    """
                    if q_val[surgery, surgeon, d, l] == 1:
                        print("surgeon={:}".format(surgeon.get_surgeon_id()))
                    """
            for r in list_room:
                ot_val[r, d, l] = ot[r, d, l].value()

            planned_surgery = [surgery for surgery in waiting_list if
                               sum(x_val[surgery, r, d, l] for r in list_room) == 1]
            planned_surgeon = [surgeon for surgeon in list_surgeon if
                               sum(q_val[surgery, surgeon, d, l] for surgery in planned_surgery) >= 1]
            dict_surgery_room = {}
            for r in list_room:
                list_surgery_in_r = [surgery for surgery in planned_surgery if x_val[surgery, r, d, l] == 1]
                dict_surgery_room[r] = list_surgery_in_r
            dict_surgery_surgeon = {}
            for surgeon in planned_surgeon:
                list_surgery_by_k = [surgery for surgery in planned_surgery if q_val[surgery, surgeon, d, l] == 1]
                dict_surgery_surgeon[surgeon] = list_surgery_by_k

            for r in list_room:
                list_surgery_in_r = dict_surgery_room[r]
                for surgery1 in list_surgery_in_r:
                    for surgery2 in list_surgery_in_r:
                        if surgery1 != surgery2:
                            y[surgery1, surgery2, r, d, l] = pulp.LpVariable(
                                "y({:},{:},{:},{:},{:})".format(surgery1.get_surgery_id(), surgery2.get_surgery_id(), r,
                                                                d, l), cat='Binary')

            for surgeon in planned_surgeon:
                list_surgery_by_k = dict_surgery_surgeon[surgeon]
                for surgery1 in list_surgery_by_k:
                    for surgery2 in list_surgery_by_k:
                        if surgery1 != surgery2:
                            z[surgery1, surgery2, surgeon, d, l] = pulp.LpVariable(
                                "z({:},{:},{:},{:},{:})".format(surgery1.get_surgery_id(), surgery2.get_surgery_id(),
                                                                surgeon.get_surgeon_id(), d, l), cat='Binary')

            for surgery in planned_surgery:
                ts[surgery, l] = pulp.LpVariable("ts({:},{:})".format(surgery.get_surgery_id(), l), lowBound=0,
                                                 cat='Continuous')

            for r in list_room:
                msR[r, d, l] = pulp.LpVariable("msR({:},{:},{:})".format(r, d, l), lowBound=0, cat='Continuous')

            for r in list_room:
                ot2[r, d, l] = pulp.LpVariable("ot2({:},{:},{:})".format(r, d, l), lowBound=0, cat='Continuous')

            for surgeon in planned_surgeon:
                tsS[surgeon, d, l] = pulp.LpVariable("tsS({:},{:},{:})".format(surgeon.get_surgeon_id(), d, l),
                                                     lowBound=0, cat='Continuous')

            for surgeon in planned_surgeon:
                msS[surgeon, d, l] = pulp.LpVariable("msS({:},{:},{:})".format(surgeon.get_surgeon_id(), d, l),
                                                     lowBound=0, cat='Continuous')

            for surgeon in planned_surgeon:
                wt[surgeon, d, l] = pulp.LpVariable("wt({:},{:},{:})".format(surgeon.get_surgeon_id(), d, l),
                                                    lowBound=0, cat='Continuous')
            operating_room_scheduling = pulp.LpProblem("ORS", pulp.LpMinimize)
            operating_room_scheduling += m6 * pulp.lpSum(ot2[r, d, l] for r in list_room) + m7 * pulp.lpSum(
                wt[surgeon, d, l] for surgeon in planned_surgeon) + m8 * pulp.lpSum(
                surgery.get_priority() * ts[surgery, l] for surgery in planned_surgery)
            for r in list_room:
                list_surgery_in_r = dict_surgery_room[r]
                for surgery1 in list_surgery_in_r:
                    for surgery2 in list_surgery_in_r:
                        if surgery1 != surgery2:
                            operating_room_scheduling += y[surgery1, surgery2, r, d, l] + y[
                                surgery2, surgery1, r, d, l] == 1, 'c1_' + str(surgery1.get_surgery_id()) + '_' + str(
                                surgery2.get_surgery_id()) + '_' + str(r)

            for surgeon in planned_surgeon:
                list_surgery_by_k = dict_surgery_surgeon[surgeon]
                for surgery1 in list_surgery_by_k:
                    for surgery2 in list_surgery_by_k:
                        if surgery1 != surgery2:
                            operating_room_scheduling += z[surgery1, surgery2, surgeon, d, l] + z[
                                surgery2, surgery1, surgeon, d, l] == 1

            for surgery in planned_surgery:
                mean = distribution_dict[surgery.get_group(), 'preparation_mean']
                variance = distribution_dict[surgery.get_group(), 'preparation_variance']
                if surgery == list_long_surgery[0]:
                    operating_room_scheduling += ts[surgery, l] == get_right_hand(alpha, mean, variance)
                else:
                    operating_room_scheduling += ts[surgery, l] >= get_right_hand(alpha, mean, variance)

            for r in list_room:
                list_surgery_in_r = dict_surgery_room[r]
                for surgery1 in list_surgery_in_r:
                    for surgery2 in list_surgery_in_r:
                        if surgery1 != surgery2:
                            g1 = surgery1.get_group()
                            g2 = surgery2.get_group()
                            mean = distribution_dict[g2, 'preparation_mean'] + distribution_dict[g1, 'surgery_mean'] \
                                + distribution_dict[g1, 'cleaning_mean']
                            variance = distribution_dict[g2, 'preparation_variance'] + distribution_dict[
                                g1, 'surgery_variance'] + distribution_dict[g1, 'cleaning_variance']
                            operating_room_scheduling += ts[surgery2, l] - ts[surgery1, l] + M * (
                                    1 - y[surgery1, surgery2, r, d, l]) >= get_right_hand(alpha, mean, variance)

            for surgeon in planned_surgeon:
                list_surgery_by_k = dict_surgery_surgeon[surgeon]
                for surgery1 in list_surgery_by_k:
                    for surgery2 in list_surgery_by_k:
                        if surgery1 != surgery2:
                            mean = distribution_dict[surgery1.get_group(), 'surgery_mean']
                            variance = distribution_dict[surgery1.get_group(), 'surgery_variance']
                            operating_room_scheduling += ts[surgery2, l] - ts[surgery1, l] - PT + M * (
                                    1 - z[surgery1, surgery2, surgeon, d, l]) >= get_right_hand(alpha, mean,
                                                                                                variance)

            for surgeon in planned_surgeon:
                for surgery in dict_surgery_surgeon[surgeon]:
                    operating_room_scheduling += tsS[surgeon, d, l] <= ts[surgery, l]

            for r in list_room:
                for surgery in dict_surgery_room[r]:
                    g = surgery.get_group()
                    mean = distribution_dict[g, 'surgery_mean'] + distribution_dict[g, 'cleaning_mean']
                    variance = distribution_dict[g, 'surgery_variance'] + distribution_dict[g, 'cleaning_variance']
                    operating_room_scheduling += msR[r, d, l] - ts[surgery, l] >= get_right_hand(alpha, mean, variance)

            for surgeon in planned_surgeon:
                for surgery in dict_surgery_surgeon[surgeon]:
                    g = surgery.get_group()
                    mean = distribution_dict[g, 'surgery_mean']
                    variance = distribution_dict[g, 'surgery_variance']
                    operating_room_scheduling += msS[surgeon, d, l] - ts[surgery, l] >= get_right_hand(alpha, mean,
                                                                                                       variance)

            for r in list_room:
                operating_room_scheduling += ot2[r, d, l] >= msR[r, d, l] - T

            for r in list_room:
                if x_val[list_long_surgery[0], r, d, l] != 1:
                    operating_room_scheduling += ot2[r, d, l] <= 120

            for surgeon in planned_surgeon:
                mean = sum(distribution_dict[s.get_group(), 'surgery_mean'] for s in dict_surgery_surgeon[surgeon])
                variance = sum(
                    distribution_dict[s.get_group(), 'surgery_variance'] for s in dict_surgery_surgeon[surgeon])
                operating_room_scheduling += msS[surgeon, d, l] - tsS[surgeon, d, l] - wt[
                    surgeon, d, l] <= get_right_hand(1 - alpha, mean, variance)

            result_status2 = operating_room_scheduling.solve(solver)
            print('ORS={:}'.format(pulp.LpStatus[result_status2]))
            if pulp.LpStatus[result_status2] == "Infeasible":
                continue
            else:
                feasibility_criteria = True
                L.append(l)
                for surgery in planned_surgery:
                    for r in list_room:
                        x_val[surgery, r, d] = x_val[surgery, r, d, l]
                    for surgeon in planned_surgeon:
                        q_val[surgery, surgeon, d] = q_val[surgery, surgeon, d, l]
                for surgery in planned_surgery:
                    ts_val[surgery] = ts[surgery, l].value()

                for r in list_room:
                    ot_val[r, d] = ot2[r, d, l].value()
                for surgeon in planned_surgeon:
                    wt_val[surgeon, d] = wt[surgeon, d, l].value()
                for surgery in planned_surgery:
                    for r in list_room:
                        if x_val[surgery, r, d] == 1:
                            scheduled_surgery.append(surgery)
                unscheduled_surgery = list(set(waiting_list) - set(scheduled_surgery))
                print(len(unscheduled_surgery))
                for surgery in unscheduled_surgery:
                    print("here")
                    ot3 = {}
                    x2 = {}
                    for r in list_room:
                        ot3[r, d] = pulp.LpVariable('ot3({:},{:})'.format(r, d), lowBound=0, cat='Continuous')
                        x2[surgery, r, d] = pulp.LpVariable('x2({:},{:},{:})'.format(surgery.get_surgery_id(), r, d),
                                                            cat='Binary')
                    q2 = {}
                    for surgeon in list_surgeon:
                        q2[surgery, surgeon, d] = pulp.LpVariable(
                            'q2({:},{:},{:})'.format(surgery.get_surgery_id(), surgeon.get_surgeon_id(), d),
                            cat='Binary')
                    y2 = {}
                    z2 = {}
                    for surgery2 in scheduled_surgery:
                        for r in list_room:
                            y2[surgery, surgery2, r, d] = pulp.LpVariable(
                                'y2({:},{:},{:},{:})'.format(surgery.get_surgery_id(), surgery2.get_surgery_id(), r, d),
                                cat='Binary')
                            y2[surgery2, surgery, r, d] = pulp.LpVariable(
                                'y2({:},{:},{:},{:})'.format(surgery2.get_surgery_id(), surgery.get_surgery_id(), r, d),
                                cat='Binary')
                        for surgeon in list_surgeon:
                            z2[surgery, surgery2, surgeon, d] = pulp.LpVariable(
                                'z2({:},{:},{:},{:})'.format(surgery.get_surgery_id(), surgery2.get_surgery_id(),
                                                             surgeon.get_surgeon_id(), d), cat='Binary')
                            z2[surgery2, surgery, surgeon, d] = pulp.LpVariable(
                                'z2({:},{:},{:},{:})'.format(surgery2.get_surgery_id(), surgery.get_surgery_id(),
                                                             surgeon.get_surgeon_id(), d), cat='Binary')
                    ts2 = {surgery: pulp.LpVariable('ts2({:})'.format(surgery.get_surgery_id()), lowBound=0,
                                                    cat='Continuous')}
                    subproblem = pulp.LpProblem('sub', pulp.LpMinimize)
                    subproblem += m6 * sum(ot3[r, d] for r in list_room) + m8 * ts2[surgery]
                    subproblem += sum(x2[surgery, r, d] for r in list_room) == 1
                    subproblem += sum(q2[surgery, surgeon, d] for surgeon in list_surgeon) == 1
                    subproblem += ts2[surgery] >= get_right_hand(alpha, surgery.get_preparation_mean(), distribution_dict[surgery.get_group(), 'preparation_variance'])
                    for r in dict_not_available_room[surgery]:
                        subproblem += x2[surgery, r, d] == 0
                    group = surgery.get_group()
                    for surgeon in list_surgeon:
                        if float(surgeon.get_dict_group()[group]) <= 0:
                            subproblem += q2[surgery, surgeon, d] == 0

                    for r in list_room:
                        for surgery2 in scheduled_surgery:
                            subproblem += 2 * (y2[surgery, surgery2, r, d] + y2[surgery2, surgery, r, d]) <= x2[
                                surgery, r, d] + x_val[surgery2, r, d]
                            subproblem += y2[surgery, surgery2, r, d] + y2[surgery2, surgery, r, d] >= x2[
                                surgery, r, d] + x_val[surgery2, r, d] - 1
                            mean1 = surgery.get_surgery_mean() + surgery.get_cleaning_mean() + surgery2.get_preparation_mean()
                            variance1 = distribution_dict[surgery.get_group(), 'surgery_variance'] + distribution_dict[surgery.get_group(), 'cleaning_variance'] + distribution_dict[surgery2.get_group(), 'preparation_variance']
                            subproblem += ts_val[surgery2] - ts2[surgery] + M * (1 - y2[surgery, surgery2, r, d]) >= get_right_hand(alpha, mean1, variance1)
                            mean2 = surgery2.get_surgery_mean() + surgery2.get_cleaning_mean() + surgery.get_preparation_mean()
                            variance2 = distribution_dict[surgery2.get_group(), 'surgery_variance'] + distribution_dict[surgery2.get_group(), 'cleaning_variance'] + distribution_dict[surgery.get_group(), 'preparation_variance']

                            subproblem += ts2[surgery] - ts_val[surgery2] + M * (1 - y2[surgery2, surgery, r, d]) >= get_right_hand(alpha, mean2, variance2)
                    for surgeon in list_surgeon:
                        for surgery2 in scheduled_surgery:
                            subproblem += 2 * (z2[surgery, surgery2, surgeon, d] + z2[surgery2, surgery, surgeon, d]) <= \
                                          q2[surgery, surgeon, d] + q_val[surgery2, surgeon, d]
                            subproblem += z2[surgery, surgery2, surgeon, d] + z2[surgery2, surgery, surgeon, d] >= q2[
                                surgery, surgeon, d] + q_val[surgery2, surgeon, d] - 1
                            mean1 = surgery.get_surgery_mean()
                            variance1 = distribution_dict[surgery.get_group(), 'surgery_variance']
                            subproblem += ts_val[surgery2] - ts2[surgery] + M * (1 - z2[surgery, surgery2, surgeon, d]) - PT >= get_right_hand(alpha, mean1, variance1)
                            mean2 = surgery2.get_surgery_mean()
                            variance2 = distribution_dict[surgery2.get_group(), 'surgery_variance']
                            subproblem += ts2[surgery] - ts_val[surgery2] - PT + M * (1 - z2[surgery2, surgery, surgeon, d]) >= get_right_hand(alpha, mean2, variance2)
                    for r in list_room:
                        mean = surgery.get_surgery_mean() + surgery.get_cleaning_mean()
                        variance = distribution_dict[surgery.get_group(), 'surgery_variance'] + distribution_dict[surgery.get_group(), 'cleaning_variance']
                        subproblem += ot3[r, d] + T - ts2[surgery] + M * (1 - x2[surgery, r, d]) >= get_right_hand(alpha, mean, variance)
                        subproblem += ot3[r, d] <= O_max

                    result_status3 = subproblem.solve(solver)
                    if pulp.LpStatus[result_status3] != 'Optimal':
                        continue
                    else:
                        for r in list_room:
                            x_val[surgery, r, d] = round(x2[surgery, r, d].value())
                        for surgeon in list_surgeon:
                            q_val[surgery, surgeon, d] = round(q2[surgery, surgeon, d].value())
                        ts_val[surgery] = ts2[surgery].value()
                        scheduled_surgery.append(surgery)

        print('スケジュールされた手術:{:}'.format(len(scheduled_surgery)))
        if len(list_surgery) == len(scheduled_surgery):
            break
        for surgery in list_surgery:
            for r in list_room:
                if surgery in waiting_list and x_val[surgery, r, d] == 1:
                    waiting_list.remove(surgery)
    end = time.time()

    rest_surgery = list(set(list_surgery) - set(scheduled_surgery))
    print("残り手術数={:}".format(len(rest_surgery)))
    print("計算時間 = {:}".format(end - start))
    print("手術数 = {:}".format(len(list_surgery)))
    print("外科医の人数 = {:}".format(len(list_surgeon)))
    objective_value = objective_function(list_surgery, list_surgeon, list_room, list_date, x_val, ot_val, wt_val,
                                         ts_val)
    print("目的関数値={:}".format(objective_value))
    over_time = sum(ot_val[r, d] for r in list_room for d in list_date)
    print("残業時間={:}".format(over_time))
    print('scheduled surgery')
    for surgery in scheduled_surgery:
        print(surgery.get_surgery_id(), surgery.get_surgery_mean(), surgery.get_group(), get_scheduled_surgery_info(surgery, list_room, list_surgeon, list_date, x_val, q_val), ts_val[surgery])
    print('unscheduled surgery')
    for surgery in rest_surgery:
        print(surgery.get_surgery_id(), np.exp(distribution_dict[surgery.get_group(), 'total_mu'] + norm.ppf(alpha) * np.sqrt(distribution_dict[surgery.get_group(), 'total_sigma'])), surgery.get_group())
    print(len(list_surgery), len(list_surgeon), len(list_room), len(list_date))
    for d in list_date:
        for r in list_room:
            if get_empty_room(r, d, list_surgery, x_val):
                print(r, d)
    create_excel_file(scheduled_surgery, list_date, list_room, ts_val, x_val)
    create_excel_file2(scheduled_surgery, list_date, list_surgeon, ts_val, q_val)
    graph.create_ganttchart(file_path, save_path)
    graph.create_surgeon_ganttchart(file_path2, save_path)


def objective_function(lst_surgery, lst_surgeon, lst_room, lst_date, x_val, ot_val, wt_val, ts_val):
    objective = sum((m1 * (d - surgery.get_release_date()) + m2 * max(d - surgery.get_due_date(), 0)) * x_val[
        surgery, r, d] * surgery.get_priority() for surgery in lst_surgery for r in lst_room for d in lst_date)
    objective += sum((m3 * (surgery.get_due_date() - surgery.get_release_date()) + m4 * (
            len(lst_date) + 1 - surgery.get_due_date())) * surgery.get_priority() * (
                             1 - sum(x_val[surgery, r, d] for r in lst_room for d in lst_date)) for surgery in
                     lst_surgery)
    objective += m5 * sum(
        surgery.get_priority() * (1 - sum(x_val[surgery, r, d] for r in lst_room for d in lst_date)) for surgery in
        lst_surgery)
    objective += m6 * sum(ot_val[r, d] for r in lst_room for d in lst_date)
    objective += m7 * sum(wt_val[surgeon, d] for surgeon in lst_surgeon for d in lst_date) + m8 * sum(
        surgery.get_priority() * ts_val[surgery] for surgery in lst_surgery)
    return objective


def sort(list_surgery, ts_val):
    if len(list_surgery) <= 1 or len(set(list_surgery)) == 1:
        return list_surgery
    else:
        pivot = np.random.choice(list_surgery)
        pivot_val = ts_val[pivot]
        left = []
        right = []
        for i in range(len(list_surgery)):
            surgery = list_surgery[i]
            start_time = ts_val[surgery]
            if start_time < pivot_val:
                left.append(surgery)
            else:
                right.append(surgery)
        return sort(left, ts_val) + sort(right, ts_val)


def get_surgeries_in_r(list_surgery, room, date, x_val, ts_val):
    lst = []
    for surgery in list_surgery:
        if (surgery, room, date) in x_val.keys() and x_val[surgery, room, date] == 1:
            lst.append(surgery)
    return sort(lst, ts_val)


def get_surgeries_by_k(list_surgery, surgeon, date, q_val, ts_val):
    lst = []
    for surgery in list_surgery:
        if (surgery, surgeon, date) in q_val.keys() and q_val[surgery, surgeon, date] == 1:
            lst.append(surgery)

    return sort(lst, ts_val)


def create_excel_file(list_surgery, list_date, list_room, ts_val, x_val):
    book = px.Workbook()
    for l in range(len(list_date)):
        d = list_date[l]
        sheet = book.worksheets[l]
        sheet.title = 'surgery_schedule' + str(d)
        length = 0
        scheduled_surgery_d = []
        for s in list_surgery:
            for r in list_room:
                if (s, r, d) in x_val.keys() and x_val[s, r, d] == 1:
                    scheduled_surgery_d.append(s)
        for i in range(len(list_room)):
            for j in range(len(scheduled_surgery_d) * 4):
                r = list_room[i]
                sheet.cell(row=j + 2, column=list_room.index(r) + 2).value = 0
        for r in list_room:
            list_surgery_in_r = get_surgeries_in_r(scheduled_surgery_d, r, d, x_val, ts_val)
            sheet.cell(row=1, column=list_room.index(r) + 2).value = "手術室" + r
            for i in range(len(list_surgery_in_r) * 4):
                surgery = list_surgery_in_r[int(i / 4)]
                if i % 4 == 0:
                    sheet.cell(row=i + 2 + length, column=1).value = "空白"
                    if i == 0:
                        sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = ts_val[
                                                                                                  surgery] - surgery.get_random_preparation()
                    else:
                        prev_surgery = list_surgery_in_r[int(i / 4 - 1)]
                        entry_time = ts_val[surgery] - surgery.get_random_preparation()
                        prev_exit_time = ts_val[
                                             prev_surgery] + prev_surgery.get_random_surgery() + prev_surgery.get_random_cleaning()
                        sheet.cell(row=i + 2 + length,
                                   column=list_room.index(r) + 2).value = entry_time - prev_exit_time
                elif i % 4 == 1:
                    sheet.cell(row=i + 2 + length, column=1).value = "準備" + str(surgery)
                    sheet.cell(row=i + 2 + length,
                               column=list_room.index(r) + 2).value = surgery.get_random_preparation()
                elif i % 4 == 2:
                    sheet.cell(row=i + 2 + length, column=1).value = "手術" + str(surgery)
                    sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = surgery.get_random_surgery()
                if i % 4 == 3:
                    sheet.cell(row=i + 2 + length, column=1).value = "清掃" + str(surgery)
                    sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = surgery.get_random_cleaning()
            length += len(list_surgery_in_r) * 4
        if l < len(list_date) - 1:
            book.create_sheet()
    book.save(file_path)
    book.close()


def create_excel_file2(lst_surgery, lst_date, lst_surgeon, ts_val, q_val):
    book = px.Workbook()
    for l in range(len(lst_date)):
        d = lst_date[l]
        sheet = book.worksheets[l]
        sheet.title = 'surgery_schedule' + str(d)
        length = 0
        scheduled_surgery_d = []
        schedule_surgeon_d = []
        for surgery in lst_surgery:
            for surgeon in lst_surgeon:
                if q_val[surgery, surgeon, d] == 1:
                    scheduled_surgery_d.append(surgery)
                    if surgeon not in schedule_surgeon_d:
                        schedule_surgeon_d.append(surgeon)
        for i in range(len(schedule_surgeon_d)):
            for j in range(len(lst_surgery) * 2):
                sheet.cell(row=j + 2, column=i + 2).value = 0
        for i in range(len(schedule_surgeon_d)):
            surgeon = schedule_surgeon_d[i]
            list_surgery_by_surgeon = get_surgeries_by_k(lst_surgery, surgeon, d, q_val, ts_val)
            sheet.cell(row=1, column=i + 2).value = "外科医" + str(surgeon.get_surgeon_id())
            for j in range(len(list_surgery_by_surgeon) * 2):
                surgery = list_surgery_by_surgeon[int(j / 2)]
                if j % 2 == 0:
                    sheet.cell(row=j + 2 + length, column=1).value = "空白"
                    if j == 0:
                        sheet.cell(row=j + 2 + length, column=i + 2).value = ts_val[surgery]
                    else:
                        prev_surgery = list_surgery_by_surgeon[int(j / 2 - 1)]
                        sheet.cell(row=j + 2 + length, column=i + 2).value = ts_val[surgery] - (ts_val[
                                                                                                    prev_surgery] + prev_surgery.get_random_surgery() + prev_surgery.get_random_cleaning())
                else:
                    sheet.cell(row=j + 2 + length, column=1).value = "手術" + surgery.get_surgery_id()
                    sheet.cell(row=j + 2 + length, column=i + 2).value = surgery.get_random_surgery()
            length += len(list_surgery_by_surgeon) * 2
        if l < len(lst_date) - 1:
            book.create_sheet()
    book.save(file_path2)
    book.close()


if __name__ == '__main__':
    main()
