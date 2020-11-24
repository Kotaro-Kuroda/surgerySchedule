import pyomo.environ as pyo
import os
import numpy as np
import time
import openpyxl as px
import graph
import copy
import instance
from scipy.stats import norm

home = os.environ['HOME']
directory_path = home + '/Documents/data-july'
path = directory_path + '/operations_with_jisseki_remake.csv'
surgeon_info_path = directory_path + '/surgeon_info.csv'
file_path = home + '/Documents/surgerySchedule/room_schedule_pyomo_heuristic1.xlsx'
file_path2 = home + '/Documents/surgerySchedule/surgeon_schedule_pyomo_heuristic1.xlsx'
surgeries_path = directory_path + '/surgeries.csv'
distribution_path = home + '/Documents/surgerySchedule/distribution.csv'
num_surgery = 10
num_date = 1
T = 480
PT = 15
M = 10 ** 5
O_max = 120
alpha = 0.95
seed = 1
surgery_instance = instance.SurgeryInstance(path, surgeon_info_path, num_surgery, num_date, seed, surgeries_path, distribution_path)
list_room, list_surgery, list_surgeon, list_date = surgery_instance.get_sets()
distribution_dict = surgery_instance.get_dict_distribution()
waiting_list = []
list_surgery_copied = copy.copy(list_surgery)
dict_surgery_id_room = surgery_instance.get_room_surgery_dict()

x_val = {}
q_val = {}
ts_val = {}
ot_val = {}

scheduled_surgery = []

dict_not_available_room = {}
for surgery in list_surgery:
    types = surgery.get_surgery_type()
    available_room = set(list_room)
    for s_type in types:
        if s_type in dict_surgery_id_room.keys():
            available_room = available_room & set(dict_surgery_id_room[s_type])
        else:
            available_room = available_room & set(list_room)
    not_available_room = set(list_room) - available_room
    dict_not_available_room[surgery] = list(not_available_room)
def aprox_norm_cdf(x):
    return 1 - 1 / np.sqrt(2 * np.pi) * pyo.exp(-x**2 / 2) / (0.226 + 0.64 * x + 0.33 * pyo.sqrt(x**2 + 3))

def get_right_hand(alpha, u, d):
    if u > 0 and d > 0:
        sigma = np.log(d / (u ** 2) + 1)
        mu = np.log(u) - sigma / 2
        return np.exp(np.sqrt(sigma) * norm.ppf(alpha) + mu)
    else:
        return 0

def get_left_hand(alpha, u, d):
    return pyo.log(u) - 1 / 2 * pyo.log(d + u ** 2) + 1 / 2 * pyo.log(u ** 2) + pyo.sqrt(pyo.log(d + u ** 2) - pyo.log(u ** 2)) * norm.ppf(alpha)
def combS_init(model):
    return ((s1, s2) for s1 in model.S for s2 in model.S if s1 != s2)
def R_init(model, s):
    return dict_not_available_room[s]

def rule1(model, s):
    return sum(model.x[s, r] for r in model.R) <= 1

def rule2(model, s):
    return sum(model.q[s, k] for k in model.K) <= 1

def rule3(model, s):
    return sum(model.x[s, r] for r in model.R) == sum(model.q[s, k] for k in model.K)

"""prod = 1
                for s in model.S:
                    prod *= (1 - model.x[s, r])"""
def rule4(model, r):
    prod = pyo.prod(1 - model.x[s, r] for s in model.S)
    mean = sum(model.x[s, r] * distribution_dict[s.get_group(), 'total_mean'] for s in model.S) + prod
    variance = sum(model.x[s, r] * distribution_dict[s.get_group(), 'total_variance'] for s in model.S) + prod / 100
    return get_left_hand(alpha, mean, variance) <= pyo.log(model.ot[r] + T)
def rule5(model, s):
    set_not_available_room = list(model.R_not_available[s])
    if len(set_not_available_room) > 0:
        return sum(model.x[s, r] for r in set_not_available_room) == 0
    else:
        return pyo.Constraint.NoConstraint

def rule6(model, r):
    return model.ot[r] <= O_max

def rule7(model, r):
    return model.p[r] >= 1 - sum(model.x[s, r] for s in model.S)

def rule8(model, s, r):
    return model.p[r] <= 1 - model.x[s, r]

def fc1(model, s1, s2, j):
    if sum(x_val[s1, r, d, j] for r in model.R) == 0 and sum(x_val[s2, r, d, j] for r in model.R) == 0 and sum(q_val[s1, k, d, j] for k in model.K) == 0 and sum(q_val[s2, k, d, j] for k in model.K) == 0:
        return pyo.Constraint.NoConstraint
    else:
        return sum(model.x[s1, r] for r in model.R if x_val[s1, r, d, j] == 1) + sum(model.x[s2, r] for r in model.R if x_val[s2, r, d, j] == 1) + sum(model.q[s1, k] for k in model.K if q_val[s1, k, d, j] == 1) + sum(model.q[s2, k] for k in model.K if q_val[s2, k, d, j] == 1) <= 3


def objective(model):
    ob = 3 * sum(B[s, r] * ((d - s.get_release_date() + max(d - s.get_due_date(), 0)) * model.x[s, r] * s.get_priority()) for s in model.S for r in model.R)
    ob += 30 * sum((s.get_due_date() - s.get_release_date() + len(list_date) + 1 - s.get_due_date()) * (1 - sum(model.x[s, r] for r in model.R)) * s.get_priority() for s in model.S)
    ob += 1000 * sum(s.get_priority() * (1 - sum(model.x[s, r] for r in model.R)) for s in model.S)
    ob += 20 * sum(model.A[r] * model.ot[r] for r in model.R)
    return ob

def rule_s1(model, s1, s2, r):
    if x_val[s1, r, d, l] == 1 and x_val[s2, r, d, l] == 1:
        return model.y[s1, s2, r] + model.y[s2, s1, r] == 1
    else:
        return pyo.Constraint.NoConstraint

def rule_s2(model, s1, s2, k):
    if q_val[s1, k, d, l] == 1 and q_val[s2, k, d, l] == 1:
        return model.z[s1, s2, k] + model.z[s2, s1, k] == 1
    else:
        return pyo.Constraint.NoConstraint

def rule_s3(model, s1, s2, r):
    g1 = s1.get_group()
    g2 = s2.get_group()
    mean = distribution_dict[g2, 'preparation_mean'] + distribution_dict[g1, 'surgery_mean'] + distribution_dict[g1, 'cleaning_mean']
    variance = distribution_dict[g2, 'preparation_variance'] + distribution_dict[g1, 'surgery_variance'] + distribution_dict[g1, 'cleaning_variance']
    return model.ts[s2] - model.ts[s1] + M * (1 - model.y[s1, s2, r]) >= get_right_hand(alpha, mean, variance)

def rule_s4(model, s1, s2, k):
    g = s1.get_group()
    mean = distribution_dict[g, 'surgery_mean']
    variance = distribution_dict[g, 'surgery_variance']
    return model.ts[s2] - model.ts[s1] - PT + M * (1 - model.z[s1, s2, k]) >= get_right_hand(alpha, mean, variance)

def rule_s5(model, s):
    g = s.get_group()
    mean = distribution_dict[g, 'preparation_mean']
    variance = distribution_dict[g, 'preparation_variance']
    if sum(x_val[s, r, d, l] for r in model.R) == 1:
        return model.ts[s] >= get_right_hand(alpha, mean, variance)
    else:
        return pyo.Constraint.NoConstraint
def rule_s6(model, s, k):
    return model.tsS[k] <= model.ts[s] + M * (1 - q_val[s, k, d, l])

def rule_s7(model, s, r):
    g = s.get_group()
    mean = distribution_dict[g, 'surgery_mean'] + distribution_dict[g, 'cleaning_mean']
    variance = distribution_dict[g, 'surgery_variance'] + distribution_dict[g, 'cleaning_variance']
    if x_val[s, r, d, l] == 1:
        return model.msR[r] - model.ts[s] >= get_right_hand(alpha, mean, variance)
    else:
        return pyo.Constraint.NoConstraint

def rule_s8(model, s, k):
    g = s.get_group()
    mean = distribution_dict[g, 'surgery_mean']
    variance = distribution_dict[g, 'surgery_variance']
    if q_val[s, k, d, l] == 1:
        return model.msS[k] - model.ts[s] + M * (1 - q_val[s, k, d, l]) >= get_right_hand(alpha, mean, variance)
    else:
        return pyo.Constraint.NoConstraint

def rule_s9(model, r):
    return model.ot[r] >= model.msR[r] - T

def rule_s10(model, r):
    return model.ot[r] <= O_max

def rule_s11(model, k):
    mean = sum(distribution_dict[s.get_group(), 'surgery_mean'] * q_val[s, k, d, l] for s in model.S)
    variance = sum(distribution_dict[s.get_group(), 'surgery_variance'] * q_val[s, k, d, l] ** 2 for s in model.S)
    return model.msS[k] - model.tsS[k] - model.wt[k] <= get_right_hand(1 - alpha, mean, variance)

def objective_s(model):
    return sum(model.ot[r] for r in model.R) + sum(model.wt[k] for k in model.K) + sum(s.get_priority() * model.ts[s] for s in model.S)
opt = pyo.SolverFactory("mindtpy")

# opt.options["limits/gap"] = 0.01

start = time.time()
for d in list_date:
    feasiblity_criteria = False
    print("date = " + str(d))
    for surgery in list_surgery_copied[:]:
        if surgery.get_release_date() <= d and surgery not in waiting_list:
            waiting_list.append(surgery)
            list_surgery_copied.remove(surgery)

    l = 0
    if len(waiting_list) == 0:
        continue
    while not feasiblity_criteria:
        l += 1
        print("loop = " + str(l))
        model = pyo.ConcreteModel()
        model.S = pyo.Set(initialize=waiting_list)
        model.R = pyo.Set(initialize=list_room)
        model.R_not_available = pyo.Set(model.S, initialize=R_init)
        model.K = pyo.Set(initialize=list_surgeon)
        model.comb_S = pyo.Set(initialize=combS_init)
        model.x = pyo.Var(model.S, model.R, domain=pyo.Binary)
        model.q = pyo.Var(model.S, model.K, domain=pyo.Binary)
        # model.p = pyo.Var(model.R, domain=pyo.Binary)
        model.ot = pyo.Var(model.R, domain=pyo.NonNegativeReals)
        A = {}
        for r in model.R:
            A[r] = 1 + 0.05 * np.random.random()
        model.A = pyo.Param(model.R, initialize=A)
        B = {}
        for s in model.S:
            for r in model.R:
                B[s, r] = 1 + 0.05 * np.random.random()
        model.B = pyo.Param(model.S, model.R, initialize=B)
        model.constrain1 = pyo.Constraint(model.S, rule=rule1)
        model.constraint2 = pyo.Constraint(model.S, rule=rule2)
        model.constraint3 = pyo.Constraint(model.S, rule=rule3)
        model.constraint5 = pyo.Constraint(model.S, rule=rule5)
        model.constraint4 = pyo.Constraint(model.R, rule=rule4)
        model.constraint6 = pyo.Constraint(model.R, rule=rule6)
        # model.Constraint7 = pyo.Constraint(model.R, rule=rule7)
        # model.Constraint8 = pyo.Constraint(model.S, model.R, rule=rule8)
        model.objective = pyo.Objective(rule=objective)
        if l == 1:
            res = opt.solve(model, mip_solver='glpk', nlp_solver='ipopt', tee=True)
            print(res.solver.termination_condition)
            if res.solver.termination_condition == 'infeasible':
                print("実行不可能")
                break
            for s in model.S:
                for r in model.R:
                    x_val[s, r, d, l] = model.x[s, r]()
                    # print(f"x[{s}, {r}] = {model.x[s, r]()}")

            for s in model.S:
                for k in model.K:
                    q_val[s, k, d, l] = model.q[s, k]()
                    # print(f"q[{s}, {k}] = {model.q[s, k]()}")
        else:
            model.fc1 = pyo.Constraint(model.comb_S, pyo.RangeSet(l - 1), rule=fc1)
            res = opt.solve(model, mip_solver='glpk', nlp_solver='ipopt', tee=True)
            print(res.solver.termination_condition)
            for s in model.S:
                for r in model.R:
                    x_val[s, r, d, l] = model.x[s, r]()

            for s in model.S:
                for k in model.K:
                    q_val[s, k, d, l] = model.q[s, k]()

        temp_scheduled_surgery = [s for s in model.S if sum(x_val[s, r, d, l] for r in model.R) == 1]
        temp_scheduled_room = [r for r in model.R if sum(x_val[s, r, d, l] for s in model.S) >= 1]
        temp_scheduled_surgeon = [k for k in model.K if sum(q_val[s, k, d, l] for s in model.S) >= 1]
        operating_room_scheduling = pyo.ConcreteModel()
        operating_room_scheduling.S = pyo.Set(initialize=waiting_list)
        operating_room_scheduling.R = pyo.Set(initialize=list_room)
        operating_room_scheduling.K = pyo.Set(initialize=list_surgeon)
        operating_room_scheduling.comb_S = pyo.Set(initialize=combS_init)
        operating_room_scheduling.y = pyo.Var(operating_room_scheduling.comb_S, operating_room_scheduling.R, domain=pyo.Binary)
        operating_room_scheduling.z = pyo.Var(operating_room_scheduling.comb_S, operating_room_scheduling.K, domain=pyo.Binary)
        operating_room_scheduling.ot = pyo.Var(operating_room_scheduling.R, domain=pyo.NonNegativeReals)
        operating_room_scheduling.ts = pyo.Var(operating_room_scheduling.S, domain=pyo.NonNegativeReals)
        operating_room_scheduling.msR = pyo.Var(operating_room_scheduling.R, domain=pyo.NonNegativeReals)
        operating_room_scheduling.tsS = pyo.Var(operating_room_scheduling.K, domain=pyo.NonNegativeReals)
        operating_room_scheduling.msS = pyo.Var(operating_room_scheduling.K, domain=pyo.NonNegativeReals)
        operating_room_scheduling.wt = pyo.Var(operating_room_scheduling.K, domain=pyo.NonNegativeReals)
        operation_room_sceduling.lam = pyo.Var(pyo.RangeSet(t), domain=pyo.NonNegativeReals)
        operating_room_scheduling.constraint1 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.R, rule=rule_s1)
        # operating_room_scheduling.constraint1_1 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.R, rule=rule_s1_2)
        operating_room_scheduling.constraint2 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.K, rule=rule_s2)
        # operating_room_scheduling.constraint2_2 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.K, rule=rule_s2_2)
        operating_room_scheduling.constraint3 = pyo.Constraint(rule=rule_s3_2)
        # operating_room_scheduling.constraint3 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.R, rule=rule_s3)
        operating_room_scheduling.constraint4 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.K, rule=rule_s4)
        operating_room_scheduling.constraint5 = pyo.Constraint(operating_room_scheduling.S, rule=rule_s5)
        operating_room_scheduling.constraint6 = pyo.Constraint(operating_room_scheduling.S, operating_room_scheduling.K, rule=rule_s6)
        operating_room_scheduling.constraint7 = pyo.Constraint(operating_room_scheduling.S, operating_room_scheduling.R, rule=rule_s7)
        operating_room_scheduling.constraint8 = pyo.Constraint(operating_room_scheduling.S, operating_room_scheduling.K, rule=rule_s8)
        operating_room_scheduling.constraint9 = pyo.Constraint(operating_room_scheduling.R, rule=rule_s9)
        operating_room_scheduling.constraint10 = pyo.Constraint(operating_room_scheduling.R, rule=rule_s10)
        operating_room_scheduling.constraint11 = pyo.Constraint(operating_room_scheduling.K, rule=rule_s11)
        operating_room_scheduling.objective = pyo.Objective(rule=objective_s)
        res = opt.solve(operating_room_scheduling, mip_solver='glpk', nlp_solver='ipopt', tee=True)
        if res.solver.termination_condition == 'infeasible':
            continue
        else:
            feasiblity_criteria = True
            for s in model.S:
                for r in model.R:
                    x_val[s, r, d] = x_val[s, r, d, l]
                    ts_val[s] = operating_room_scheduling.ts[s]()
                for k in model.K:
                    q_val[s, k, d] = q_val[s, k, d, l]

            for r in model.R:
                ot_val[r, d] = operating_room_scheduling.ot[r]()
            for s in model.S:
                for r in model.R:
                    if x_val[s, r, d] == 1:
                        waiting_list.remove(s)
                        scheduled_surgery.append(s)
end = time.time()
print("計算時間={:}".format(end - start))
print("残業時間={:}".format(sum(ot_val[r, d] for r in list_room for d in list_date if (r, d) in ot_val.keys())))
print("(S, K, R) = ({:}, {:}, {:})".format(len(list_surgery), len(list_surgeon), len(list_room)))
print(len(scheduled_surgery))
print("計算終了")
def sort(list_surgery, ts_val):
    if len(list_surgery) <= 1:
        return list_surgery
    pivot = np.random.choice(list_surgery)
    pivot_val = ts_val[pivot]
    left = []
    right = []
    for i in range(len(list_surgery)):
        surgery = list_surgery[i]
        start_time = ts_val[surgery]
        if start_time < pivot_val:
            left.append(surgery)
        else:
            right.append(surgery)
    return sort(left, ts_val) + sort(right, ts_val)


def get_surgeries_in_r(list_surgery, room, date, x_val, ts_val):
    lst = []
    for surgery in list_surgery:
        if (surgery, room, date) in x_val.keys() and x_val[surgery, room, date] == 1:
            lst.append(surgery)
    return sort(lst, ts_val)


def get_surgeries_by_k(list_surgery, surgeon, date, q_val, ts_val):
    lst = []
    for surgery in list_surgery:
        if (surgery, surgeon, date) in q_val.keys() and q_val[surgery, surgeon, date] == 1:
            lst.append(surgery)

    return sort(lst, ts_val)


def create_excel_file(list_surgery, list_date, list_room, ts_val, x_val):
    book = px.Workbook()
    for l in range(len(list_date)):
        print(l)
        d = list_date[l]
        sheet = book.worksheets[l]
        sheet.title = 'surgery_schedule' + str(d)
        length = 0
        scheduled_surgery_d = []
        for s in list_surgery:
            for r in list_room:
                if (s, r, d) in x_val.keys() and x_val[s, r, d] == 1:
                    scheduled_surgery_d.append(s)
        for i in range(len(list_room)):
            for j in range(len(scheduled_surgery_d) * 4):
                r = list_room[i]
                sheet.cell(row=j + 2, column=list_room.index(r) + 2).value = 0
        for r in list_room:
            list_surgery_in_r = get_surgeries_in_r(scheduled_surgery_d, r, d, x_val, ts_val)
            sheet.cell(row=1, column=list_room.index(r) + 2).value = "手術室" + r
            for i in range(len(list_surgery_in_r) * 4):
                surgery = list_surgery_in_r[int(i / 4)]
                if i % 4 == 0:
                    sheet.cell(row=i + 2 + length, column=1).value = "空白"
                    if i == 0:
                        sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = ts_val[surgery] - surgery.get_preparation_time()
                    else:
                        prev_surgery = list_surgery_in_r[int(i / 4 - 1)]
                        entry_time = ts_val[surgery] - surgery.get_preparation_time()
                        prev_exit_time = ts_val[prev_surgery] + prev_surgery.get_surgery_time() + prev_surgery.get_cleaning_time()
                        sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = entry_time - prev_exit_time
                elif i % 4 == 1:
                    sheet.cell(row=i + 2 + length, column=1).value = "準備" + str(surgery)
                    sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = surgery.get_preparation_time()
                elif i % 4 == 2:
                    sheet.cell(row=i + 2 + length, column=1).value = "手術" + str(surgery)
                    sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = surgery.get_surgery_time()
                if i % 4 == 3:
                    sheet.cell(row=i + 2 + length, column=1).value = "清掃" + str(surgery)
                    sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = surgery.get_cleaning_time()
            length += len(list_surgery_in_r) * 4
        if l < len(list_date) - 1:
            print(l)
            book.create_sheet()
    book.save(file_path)
    book.close()

def create_excel_file2(list_surgery, list_date, list_surgeon, ts_val, x_val):
    book = px.Workbook()
    for l in range(len(list_date)):
        d = list_date[l]
        sheet = book.worksheets[l]
        sheet.title = 'surgery_schedule' + str(d)
        length = 0
        scheduled_surgery_d = []
        for s in list_surgery:
            for k in list_surgeon:
                if (s, k, d) in q_val.keys() and q_val[s, k, d] == 1:
                    scheduled_surgery_d.append(s)
        for i in range(len(list_surgeon)):
            for j in range(len(list_surgery) * 2):
                sheet.cell(row=j + 2, column=i + 2).value = 0
        for i in range(len(list_surgeon)):
            surgeon = list_surgeon[i]
            list_surgery_by_surgeon = get_surgeries_by_k(list_surgery, surgeon, d, q_val, ts_val)
            sheet.cell(row=1, column=i + 2).value = "外科医" + str(surgeon.get_surgeon_id())
            for j in range(len(list_surgery_by_surgeon) * 2):
                surgery = list_surgery_by_surgeon[int(j / 2)]
                if j % 2 == 0:
                    sheet.cell(row=j + 2 + length, column=1).value = "空白"
                    if j == 0:
                        sheet.cell(row=j + 2 + length, column=i + 2).value = ts_val[surgery]
                    else:
                        prev_surgery = list_surgery_by_surgeon[int(j / 2 - 1)]
                        sheet.cell(row=j + 2 + length, column=i + 2).value = ts_val[surgery] - (ts_val[prev_surgery] + prev_surgery.get_surgery_time() + prev_surgery.get_cleaning_time())
                else:
                    sheet.cell(row=j + 2 + length, column=1).value = "手術" + surgery.get_surgery_id()
                    sheet.cell(row=j + 2 + length, column=i + 2).value = surgery.get_surgery_time()
            length += len(list_surgery_by_surgeon) * 2
        if l < len(list_date) - 1:
            print(l)
            book.create_sheet()
    book.save(file_path2)
    book.close()

create_excel_file(scheduled_surgery, list_date, list_room, ts_val, x_val)
create_excel_file2(scheduled_surgery, list_date, list_surgeon, ts_val, x_val)
graph.create_ganttchart(file_path)
graph.create_surgeon_ganttchart(file_path2)
