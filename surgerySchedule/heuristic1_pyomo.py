import pyomo.environ as pyo
import os
import numpy as np
import time
import openpyxl as px
import graph
import copy
import instance
home = os.environ['HOME']
directory_path = home + '/Documents/data-july'
path = directory_path + '/operations_with_jisseki_remake.csv'
surgeon_info_path = directory_path + '/surgeon_info.csv'
file_path = home + '/Documents/surgerySchedule/room_schedule_pyomo_heuristic1.xlsx'
file_path2 = home + '/Documents/surgerySchedule/surgeon_schedule_pyomo_heuristic1.xlsx'
surgeries_path = directory_path + '/surgeries.csv'
distribution_path = home + '/Documents/surgerySchedule/distribution.csv'
num_surgery = 10
num_date = 1
T = 480
PT = 15
M = 10 ** 5
O_max = 120
surgery_instance = instance.SurgeryInstance(path, surgeon_info_path, num_surgery, num_date, 10, surgeries_path, distribution_path)
list_room, list_surgery, list_surgeon, list_date, list_group = surgery_instance.get_sets()
dict_surgery_id_room = surgery_instance.get_room_surgery_dict()
waiting_list = []
list_surgery_copied = copy.copy(list_surgery)

x_val = {}
q_val = {}
ts_val = {}
ot_val = {}
for surgery in list_surgery:
    for r in list_room:
        for d in list_date:
            x_val[surgery, r, d] = 0
for surgery in list_surgery:
    for surgeon in list_surgeon:
        for d in list_date:
            q_val[surgery, surgeon, d] = 0
scheduled_surgery = []

dict_not_available_room = {}
for surgery in list_surgery:
    types = surgery.get_surgery_type()
    available_room = set(list_room)
    for s_type in types:
        if s_type in dict_surgery_id_room.keys():
            available_room = available_room & set(dict_surgery_id_room[s_type])
        else:
            available_room = available_room & set(list_room)
    not_available_room = set(list_room) - available_room
    dict_not_available_room[surgery] = list(not_available_room)


def combS_init(model):
    return ((s1, s2) for s1 in model.S for s2 in model.S if s1 != s2)

def R_init(model, s):
    return dict_not_available_room[s]

def rule1(model, s):
    return sum(model.x[s, r] for r in model.R) <= 1

def rule2(model, s):
    return sum(model.q[s, k] for k in model.K) <= 1

def rule3(model, s):
    return sum(model.x[s, r] for r in model.R) == sum(model.q[s, k] for k in model.K)

def rule4(model, r):
    return sum(s.get_total_time() * model.x[s, r] for s in model.S) - T <= model.ot[r]

def rule5(model, s):
    set_not_available_room = list(model.R_not_available[s])
    if len(set_not_available_room) > 0:
        return sum(model.x[s, r] for r in set_not_available_room) == 0
    else:
        return sum(model.x[s, r] for r in model.R) <= 1

def rule6(model, k):
    return model.ost[k] >= sum(model.q[s, k] for s in model.S) - len(model.S) / len(model.K)

def rule7(model, k):
    return -model.ost[k] <= sum(model.q[s, k] for s in model.S) - len(model.S) / len(model.K)

def fc1(model, s1, s2, j):
    print(s1, s2, j)
    print(sum(model.x[s1, r] for r in model.R if x_val[s1, r, d, j] == 1) + sum(model.x[s2, r] for r in model.R if x_val[s2, r, d, j] == 1) + sum(model.q[s1, k] for k in model.K if q_val[s1, k, d, j] == 1) + sum(model.q[s2, k] for k in model.K if q_val[s2, k, d, j] == 1) <= 3)
    print(sum(x_val[s1, r, d, j] for r in model.R), sum(x_val[s2, r, d, j] for r in model.R), sum(q_val[s1, k, d, j] for k in model.K), sum(q_val[s2, k, d, j] for k in model.K))
    if sum(x_val[s1, r, d, j] for r in model.R) == 0 and sum(x_val[s2, r, d, j] for r in model.R) == 0 and sum(q_val[s1, k, d, j] for k in model.K) == 0 and sum(q_val[s2, k, d, j] for k in model.K) == 0:
        return pyo.Constraint.NoConstraint
    else:
        return sum(model.x[s1, r] for r in model.R if x_val[s1, r, d, j] == 1) + sum(model.x[s2, r] for r in model.R if x_val[s2, r, d, j] == 1) + sum(model.q[s1, k] for k in model.K if q_val[s1, k, d, j] == 1) + sum(model.q[s2, k] for k in model.K if q_val[s2, k, d, j] == 1) <= 3


def objective(model):
    ob = sum((d - s.get_release_date() + max(d - s.get_due_date(), 0)) * model.x[s, r] * s.get_priority() for s in model.S for r in model.R)
    ob += sum((s.get_due_date() - s.get_release_date() + len(list_date) + 1 - s.get_due_date()) * (1 - sum(model.x[s, r] for r in model.R)) * s.get_priority() for s in model.S)
    ob += 1000 * sum(s.get_priority() * (1 - sum(model.x[s, r] for r in model.R)) for s in model.S)
    ob += sum(model.ot[r] for r in model.R)
    ob += sum(model.ost[k] for k in model.K)
    return ob

def rule_s1(model, s1, s2, r):
    if x_val[s1, r, d, l] == 1 and x_val[s2, r, d, l] == 1:
        return model.y[s1, s2, r] + model.y[s2, s1, r] == 1
    else:
        return model.y[s1, s2, r] + model.y[s2, s1, r] == 0

"""def rule_s1_2(model, s1, s2, r):
    return model.y[s1, s2, r] + model.y[s2, s1, r] >= x_val[s1, r, d, l] + x_val[s2, r, d, l] - 1
"""
def rule_s2(model, s1, s2, k):
    if q_val[s1, k, d, l] == 1 and q_val[s2, k, d, l] == 1:
        return model.z[s1, s2, k] + model.z[s2, s1, k] == 1
    else:
        return model.z[s1, s2, k] + model.z[s2, s1, k] == 0
"""
def rule_s2_2(model, s1, s2, k):
    return model.z[s1, s2, k] + model.z[s2, s1, k] >= q_val[s1, k, d, l] + q_val[s2, k, d, l] - 1
"""
def rule_s3(model, s1, s2, r):
    return model.ts[s2] - s2.get_preparation_time() >= model.ts[s1] + s1.get_surgery_time() + s1.get_cleaning_time() - M * (1 - model.y[s1, s2, r])

def rule_s4(model, s1, s2, k):
    return model.ts[s2] >= model.ts[s1] + s1.get_surgery_time() + PT - M * (1 - model.z[s1, s2, k])

def rule_s5(model, s):
    return model.ts[s] >= s.get_preparation_time() - M * (1 - sum(x_val[s, r, d, l] for r in model.R))

def rule_s6(model, s, k):
    return model.tsS[k] <= model.ts[s] + M * (1 - q_val[s, k, d, l])

def rule_s7(model, s, r):
    return model.msR[r] >= model.ts[s] + s.get_surgery_time() + s.get_cleaning_time() - M * (1 - x_val[s, r, d, l])

def rule_s8(model, s, k):
    return model.msS[k] >= model.ts[s] + s.get_surgery_time() - M * (1 - q_val[s, k, d, l])

def rule_s9(model, r):
    return model.ot[r] >= model.msR[r] - T

def rule_s10(model, r):
    return model.ot[r] <= O_max

def rule_s11(model, k):
    return model.wt[k] >= model.msS[k] - model.tsS[k] - sum(s.get_surgery_time() * q_val[s, k, d, l] for s in model.S)

def objective_s(model):
    return 10 * sum(model.ot[r] for r in model.R) + sum(model.wt[k] for k in model.K) + 3 * sum(s.get_priority() * model.ts[s] for s in model.S)
opt = pyo.SolverFactory("cplex")

start = time.time()
for d in list_date:
    feasiblity_criteria = False
    print("date = " + str(d))
    for surgery in list_surgery_copied[:]:
        if surgery.get_release_date() <= d and surgery not in waiting_list:
            waiting_list.append(surgery)
            list_surgery_copied.remove(surgery)

    l = 0
    while not feasiblity_criteria:
        l += 1
        print("loop = " + str(l))
        model = pyo.ConcreteModel()
        model.S = pyo.Set(initialize=waiting_list)
        model.R = pyo.Set(initialize=list_room)
        model.K = pyo.Set(initialize=list_surgeon)
        model.comb_S = pyo.Set(initialize=combS_init)
        model.R_not_available = pyo.Set(model.S, initialize=R_init)
        model.x = pyo.Var(model.S, model.R, domain=pyo.Binary)
        model.q = pyo.Var(model.S, model.K, domain=pyo.Binary)
        model.ot = pyo.Var(model.R, domain=pyo.NonNegativeReals)
        model.ost = pyo.Var(model.K, domain=pyo.NonNegativeReals)

        model.constrain1 = pyo.Constraint(model.S, rule=rule1)
        model.constraint2 = pyo.Constraint(model.S, rule=rule2)
        model.constraint3 = pyo.Constraint(model.S, rule=rule3)
        model.constraint4 = pyo.Constraint(model.R, rule=rule4)
        model.constraint5 = pyo.Constraint(model.S, rule=rule5)
        model.constraint6 = pyo.Constraint(model.K, rule=rule6)
        model.constraint7 = pyo.Constraint(model.K, rule=rule7)

        model.objective = pyo.Objective(rule=objective)
        if l == 1:
            res = opt.solve(model)
            print(res.solver.termination_condition)
            for s in model.S:
                for r in model.R:
                    x_val[s, r, d, l] = round(model.x[s, r]())

            for s in model.S:
                for k in model.K:
                    q_val[s, k, d, l] = round(model.q[s, k]())
        else:
            model.L = pyo.Set(initialize=pyo.RangeSet(l - 1))
            model.fc1 = pyo.Constraint(model.comb_S, model.L, rule=fc1)
            res = opt.solve(model)
            for s in model.S:
                for r in model.R:
                    x_val[s, r, d, l] = round(model.x[s, r]())

            for s in model.S:
                for k in model.K:
                    q_val[s, k, d, l] = round(model.q[s, k]())

        operating_room_scheduling = pyo.ConcreteModel()
        operating_room_scheduling.S = pyo.Set(initialize=waiting_list)
        operating_room_scheduling.R = pyo.Set(initialize=list_room)
        operating_room_scheduling.K = pyo.Set(initialize=list_surgeon)
        operating_room_scheduling.comb_S = pyo.Set(initialize=combS_init)
        operating_room_scheduling.y = pyo.Var(operating_room_scheduling.comb_S, operating_room_scheduling.R, domain=pyo.Binary)
        operating_room_scheduling.z = pyo.Var(operating_room_scheduling.comb_S, operating_room_scheduling.K, domain=pyo.Binary)
        operating_room_scheduling.ot = pyo.Var(operating_room_scheduling.R, domain=pyo.NonNegativeReals)
        operating_room_scheduling.ts = pyo.Var(operating_room_scheduling.S, domain=pyo.NonNegativeReals)
        operating_room_scheduling.msR = pyo.Var(operating_room_scheduling.R, domain=pyo.NonNegativeReals)
        operating_room_scheduling.tsS = pyo.Var(operating_room_scheduling.K, domain=pyo.NonNegativeReals)
        operating_room_scheduling.msS = pyo.Var(operating_room_scheduling.K, domain=pyo.NonNegativeReals)
        operating_room_scheduling.wt = pyo.Var(operating_room_scheduling.K, domain=pyo.NonNegativeReals)

        operating_room_scheduling.constraint1 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.R, rule=rule_s1)
        # operating_room_scheduling.constraint1_1 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.R, rule=rule_s1_2)
        operating_room_scheduling.constraint2 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.K, rule=rule_s2)
        # operating_room_scheduling.constraint2_2 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.K, rule=rule_s2_2)
        operating_room_scheduling.constraint3 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.R, rule=rule_s3)
        operating_room_scheduling.constraint4 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.K, rule=rule_s4)
        operating_room_scheduling.constraint5 = pyo.Constraint(operating_room_scheduling.S, rule=rule_s5)
        operating_room_scheduling.constraint6 = pyo.Constraint(operating_room_scheduling.S, operating_room_scheduling.K, rule=rule_s6)
        operating_room_scheduling.constraint7 = pyo.Constraint(operating_room_scheduling.S, operating_room_scheduling.R, rule=rule_s7)
        operating_room_scheduling.constraint8 = pyo.Constraint(operating_room_scheduling.S, operating_room_scheduling.K, rule=rule_s8)
        operating_room_scheduling.constraint9 = pyo.Constraint(operating_room_scheduling.R, rule=rule_s9)
        operating_room_scheduling.constraint10 = pyo.Constraint(operating_room_scheduling.R, rule=rule_s10)
        operating_room_scheduling.constraint11 = pyo.Constraint(operating_room_scheduling.K, rule=rule_s11)
        operating_room_scheduling.objective = pyo.Objective(rule=objective_s)
        res = opt.solve(operating_room_scheduling)
        if res.solver.termination_condition == 'infeasible':
            continue
        else:
            feasiblity_criteria = True
            for s in model.S:
                for r in model.R:
                    x_val[s, r, d] = x_val[s, r, d, l]
                    ts_val[s] = operating_room_scheduling.ts[s]()
                for k in model.K:
                    q_val[s, k, d] = q_val[s, k, d, l]

            for r in model.R:
                ot_val[r, d] = operating_room_scheduling.ot[r]()
            for s in model.S:
                for r in model.R:
                    if x_val[s, r, d] == 1:
                        waiting_list.remove(s)
                        scheduled_surgery.append(s)
end = time.time()
print("計算時間={:}".format(end - start))
print("残業時間={:}".format(sum(ot_val[r, d] for r in list_room for d in list_date)))
print("(S, K, R) = ({:}, {:}, {:})".format(len(list_surgery), len(list_surgeon), len(list_room)))
print(len(scheduled_surgery))
print("計算終了")
def sort(list_surgery, ts_val):
    if len(list_surgery) <= 1:
        return list_surgery
    pivot = np.random.choice(list_surgery)
    pivot_val = ts_val[pivot]
    left = []
    right = []
    for i in range(len(list_surgery)):
        surgery = list_surgery[i]
        start_time = ts_val[surgery]
        if start_time < pivot_val:
            left.append(surgery)
        else:
            right.append(surgery)
    return sort(left, ts_val) + sort(right, ts_val)


def get_surgeries_in_r(list_surgery, room, date, x_val, ts_val):
    lst = []
    for surgery in list_surgery:
        if x_val[surgery, room, date] == 1:
            lst.append(surgery)
    return sort(lst, ts_val)


def get_surgeries_by_k(list_surgery, surgeon, date, q_val, ts_val):
    lst = []
    for surgery in list_surgery:
        if q_val[surgery, surgeon, date] == 1:
            lst.append(surgery)

    return sort(lst, ts_val)


def create_excel_file(list_surgery, list_date, list_room, ts_val, x_val):
    book = px.Workbook()
    for l in range(len(list_date)):
        d = list_date[l]
        sheet = book.worksheets[l]
        sheet.title = 'surgery_schedule' + str(d)
        length = 0
        scheduled_surgery_d = []
        for s in list_surgery:
            for r in list_room:
                if (s, r, d) in x_val.keys() and x_val[s, r, d] == 1:
                    scheduled_surgery_d.append(s)
        for i in range(len(list_room)):
            for j in range(len(scheduled_surgery_d) * 4):
                r = list_room[i]
                sheet.cell(row=j + 2, column=list_room.index(r) + 2).value = 0
        for r in list_room:
            list_surgery_in_r = get_surgeries_in_r(scheduled_surgery_d, r, d, x_val, ts_val)
            sheet.cell(row=1, column=list_room.index(r) + 2).value = "手術室" + r
            for i in range(len(list_surgery_in_r) * 4):
                surgery = list_surgery_in_r[int(i / 4)]
                if i % 4 == 0:
                    sheet.cell(row=i + 2 + length, column=1).value = "空白"
                    if i == 0:
                        sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = ts_val[surgery] - surgery.get_preparation_time()
                    else:
                        prev_surgery = list_surgery_in_r[int(i / 4 - 1)]
                        entry_time = ts_val[surgery] - surgery.get_preparation_time()
                        prev_exit_time = ts_val[prev_surgery] + prev_surgery.get_surgery_time() + prev_surgery.get_cleaning_time()
                        sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = entry_time - prev_exit_time
                elif i % 4 == 1:
                    sheet.cell(row=i + 2 + length, column=1).value = "準備" + str(surgery)
                    sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = surgery.get_preparation_time()
                elif i % 4 == 2:
                    sheet.cell(row=i + 2 + length, column=1).value = "手術" + str(surgery)
                    sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = surgery.get_surgery_time()
                if i % 4 == 3:
                    sheet.cell(row=i + 2 + length, column=1).value = "清掃" + str(surgery)
                    sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = surgery.get_cleaning_time()
            length += len(list_surgery_in_r) * 4
        if l < len(list_date) - 1:
            print(l)
            book.create_sheet()
    book.save(file_path)
    book.close()

def create_excel_file2(list_surgery, list_date, list_surgeon, ts_val, q_val):
    book = px.Workbook()
    for l in range(len(list_date)):
        d = list_date[l]
        sheet = book.worksheets[l]
        sheet.title = 'surgery_schedule' + str(d)
        length = 0
        scheduled_surgery_d = []
        schedule_surgeon_d = []
        for surgery in list_surgery:
            for surgeon in list_surgeon:
                if q_val[surgery, surgeon, d] == 1:
                    scheduled_surgery_d.append(surgery)
                    if surgeon not in schedule_surgeon_d:
                        schedule_surgeon_d.append(surgeon)
        for i in range(len(schedule_surgeon_d)):
            for j in range(len(list_surgery) * 2):
                sheet.cell(row=j + 2, column=i + 2).value = 0
        for i in range(len(schedule_surgeon_d)):
            surgeon = schedule_surgeon_d[i]
            list_surgery_by_surgeon = get_surgeries_by_k(list_surgery, surgeon, d, q_val, ts_val)
            sheet.cell(row=1, column=i + 2).value = "外科医" + str(surgeon.get_surgeon_id())
            for j in range(len(list_surgery_by_surgeon) * 2):
                surgery = list_surgery_by_surgeon[int(j / 2)]
                if j % 2 == 0:
                    sheet.cell(row=j + 2 + length, column=1).value = "空白"
                    if j == 0:
                        sheet.cell(row=j + 2 + length, column=i + 2).value = ts_val[surgery]
                    else:
                        prev_surgery = list_surgery_by_surgeon[int(j / 2 - 1)]
                        sheet.cell(row=j + 2 + length, column=i + 2).value = ts_val[surgery] - (ts_val[prev_surgery] + prev_surgery.get_surgery_time() + prev_surgery.get_cleaning_time())
                else:
                    sheet.cell(row=j + 2 + length, column=1).value = "手術" + surgery.get_surgery_id()
                    sheet.cell(row=j + 2 + length, column=i + 2).value = surgery.get_surgery_time()
            length += len(list_surgery_by_surgeon) * 2
        if l < len(list_date) - 1:
            book.create_sheet()
    book.save(file_path2)
    book.close()
create_excel_file(scheduled_surgery, list_date, list_room, ts_val, x_val)
create_excel_file2(scheduled_surgery, list_date, list_surgeon, ts_val, q_val)
graph.create_ganttchart(file_path)
graph.create_surgeon_ganttchart(file_path2)
