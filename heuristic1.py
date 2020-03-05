import copy
import csv
import os
import time
import graph
import openpyxl
import pulp
import create_instance
# 手術クラス
home = os.environ['HOME']
operations = []
with open(home + '/Documents/data-july/operations.csv') as f:
    reader = csv.reader(f)
    operations = [row for row in reader]

url = home + '/Documents/solution/solution.xlsx'
url2 = home + '/Documents/solution/solution2.xlsx'

# 手術室の集合
R = sorted(list(set([int(operations[i][5]) for i in range(1, len(operations))])))
R = [str(i) for i in R]
G = list(set([operations[i][6] for i in range(1, len(operations))]))


num_surgery = 100
num_date = 7
seed = 1
list_surgery, list_surgeon, D = create_instance.create_model(num_surgery, num_date, seed)
num_int_var = 0
num_con_var = 0
num_constraint = 0

initial = time.time()

dict_surgery_group = {}
for g in G:
    list_surgery_group = []
    for surgery in list_surgery:
        if g == surgery.get_group():
            list_surgery_group.append(surgery)
    dict_surgery_group[g] = list_surgery_group

dict_surgeon_group = {}
for g in G:
    list_surgeon_group = []
    for surgeon in list_surgeon:
        if g == surgeon.get_group():
            list_surgeon_group.append(surgeon)
    dict_surgeon_group[g] = list_surgeon_group

dict_room_group = {}
for g in G:
    list_room_group = []
    for i in range(1, len(operations)):
        if g == operations[i][6]:
            list_room_group.append(operations[i][5])
    list_room_group = list(set(list_room_group))
    dict_room_group[g] = list_room_group

dict_room_group_prob = {}
for g in G:
    for r in dict_room_group[g]:
        dict_room_group_prob[g, r] = sum(1 for i in range(1, len(operations)) if operations[i][6] == g and operations[i][5] == r) / sum(1 for i in range(1, len(operations)) if operations[i][6] == g)

TS = {surgery.get_surgery_id(): surgery.get_surgery_time() for surgery in list_surgery}

TP = {surgery.get_surgery_id(): surgery.get_preparation_time() for surgery in list_surgery}

TC = {surgery.get_surgery_id(): surgery.get_cleaning_time() for surgery in list_surgery}

A = {surgery.get_surgery_id(): surgery.get_release_date() for surgery in list_surgery}

Due = {surgery.get_surgery_id(): surgery.get_due_date() for surgery in list_surgery}

U = {surgery.get_surgery_id(): surgery.get_priority() for surgery in list_surgery}

T = 480
PT = 15
M = 99999
ope_time = 9
O_max = 120

waiting_list = []

solver = pulp.solvers.CPLEX_CMD()

def sort(lst, l):
    for i in range(len(lst) - 1):
        for j in range(i + 1, len(lst)):
            if ts[lst[i].get_surgery_id(), l].value() > ts[lst[j].get_surgery_id(), l].value():
                tmp = lst[i]
                lst[i] = lst[j]
                lst[j] = tmp
    return lst

def classify_room_date(lst, r, d, l):
    lst_result = []
    for surgery in lst:
        if x[surgery.get_surgery_id(), r, d, l].value() == 1:
            lst_result.append(surgery)
    return sort(lst_result, l)

def classify_surgeon_date(list_surgery, surgeon, d, l):
    list_result = []
    for surgery in list_surgery:
        if q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d, l].value() == 1:
            list_result.append(surgery)
    return sort(list_result, l)


def to_time(num):
    hour = int(num / 60)
    minute = int(num % 60)
    if hour < 10:
        hour = "0" + str(hour)
    if minute < 10:
        minute = "0" + str(minute)
    return str(hour) + ":" + str(minute)

def to_time_of_day(num):
    hour = int(num / 60) + 9
    minute = int(num % 60)
    if hour < 10:
        hour = "0" + str(hour)
    if minute < 10:
        minute = "0" + str(minute)
    return str(hour) + ":" + str(minute)

def get_total_time(surgery):
    return TP[surgery.get_surgery_id()] + TS[surgery.get_surgery_id()] + TC[surgery.get_surgery_id()]

def get_entry_time(surgery):
    return ts_val[surgery.get_surgery_id()] - TP[surgery.get_surgery_id()]

def get_end_time(surgery):
    return ts_val[surgery.get_surgery_id()] + TS[surgery.get_surgery_id()]

def get_exit_time(surgery):
    return get_end_time(surgery) + TC[surgery.get_surgery_id()]

def create_excel_file(sheet, list_surgery, list_room, d, l):
    length = 0
    for i in range(len(list_room)):
        for j in range(len(list_surgery) * 4):
            sheet.cell(row=j + 2, column=int(list_room[i]) + 1).value = 0
    for r in list_room:
        list_surgery_in_r = classify_room_date(list_surgery, r, d, l)
        sheet.cell(row=1, column=int(r) + 1).value = "手術室" + r
        for i in range(len(list_surgery_in_r) * 4):
            surgery = list_surgery_in_r[int(i / 4)]
            if i % 4 == 0:
                sheet.cell(row=i + 2 + length, column=1).value = "空白"
                if i == 0:
                    sheet.cell(row=i + 2 + length, column=int(r) + 1).value = ts_val[surgery.get_surgery_id()] - TP[surgery.get_surgery_id()]
                else:
                    prev_surgery = list_surgery_in_r[int(i / 4 - 1)]
                    sheet.cell(row=i + 2 + length, column=int(r) + 1).value = get_entry_time(surgery) - get_exit_time(prev_surgery)
            elif i % 4 == 1:
                sheet.cell(row=i + 2 + length, column=1).value = "準備" + str(surgery.get_surgery_id())
                sheet.cell(row=i + 2 + length, column=int(r) + 1).value = TP[surgery.get_surgery_id()]
            elif i % 4 == 2:
                sheet.cell(row=i + 2 + length, column=1).value = "手術" + str(surgery.get_surgery_id())
                sheet.cell(row=i + 2 + length, column=int(r) + 1).value = TS[surgery.get_surgery_id()]
            if i % 4 == 3:
                sheet.cell(row=i + 2 + length, column=1).value = "清掃" + str(surgery.get_surgery_id())
                sheet.cell(row=i + 2 + length, column=int(r) + 1).value = TC[surgery.get_surgery_id()]
        length += len(list_surgery_in_r) * 4

def create_excel_file2(sheet, list_surgery, list_surgeon, d, l):
    length = 0
    for i in range(len(list_surgeon)):
        for j in range(len(list_surgery) * 2):
            sheet.cell(row=j + 2, column=i + 2).value = 0
    for i in range(len(list_surgeon)):
        surgeon = list_surgeon[i]
        list_surgery_by_surgeon = classify_surgeon_date(list_surgery, surgeon, d, l)
        sheet.cell(row=1, column=i + 2).value = "外科医" + str(surgeon.get_surgeon_id())
        for j in range(len(list_surgery_by_surgeon) * 2):
            surgery = list_surgery_by_surgeon[int(j / 2)]
            if j % 2 == 0:
                sheet.cell(row=j + 2 + length, column=1).value = "空白"
                if j == 0:
                    sheet.cell(row=j + 2 + length, column=i + 2).value = ts_val[surgery.get_surgery_id()]
                else:
                    prev_surgery = list_surgery_by_surgeon[int(j / 2 - 1)]
                    sheet.cell(row=j + 2 + length, column=i + 2).value = ts_val[surgery.get_surgery_id()] - get_end_time(prev_surgery)
            else:
                sheet.cell(row=j + 2 + length, column=1).value = "手術" + surgery.get_surgery_id()
                sheet.cell(row=j + 2 + length, column=i + 2).value = TS[surgery.get_surgery_id()]
        length += len(list_surgery_by_surgeon) * 2

num = 0
start = time.time()
list_surgery_copied = copy.copy(list_surgery)
L = []
x_val = {}
for surgery in list_surgery:
    for r in R:
        for d in D:
            x_val[surgery.get_surgery_id(), r, d] = 0
ot_val = {}
for r in R:
    for d in D:
        ot_val[r, d] = 0

n_val = {}
for surgeon in list_surgeon:
    for d in D:
        n_val[surgeon.get_surgeon_id(), d] = 0

wt_val = {}
for surgeon in list_surgeon:
    for d in D:
        wt_val[surgeon.get_surgeon_id(), d] = 0

ts_val = {}
for surgery in list_surgery:
    ts_val[surgery.get_surgery_id()] = 0

x = {}
q = {}
y = {}
z = {}
n = {}
ts = {}
msR = {}
ot = {}
tsS = {}
msS = {}
wt = {}
ost = {}
ns = {}

book = openpyxl.Workbook()
book2 = openpyxl.Workbook()
for d in D:
    print("day={:}".format(d))
    feasiblity_criteria = False
    if len(list_surgery_copied) == 0:
        break
    for surgery in list_surgery_copied[:]:
        if surgery.get_release_date() <= d:
            if surgery not in waiting_list:
                waiting_list.append(surgery)
                list_surgery_copied.remove(surgery)

    l = 0
    while not feasiblity_criteria:
        l += 1
        for surgery in waiting_list:
            for r in R:
                num_int_var += 1
                x[surgery.get_surgery_id(), r, d, l] = pulp.LpVariable("x({:},{:},{:},{:})".format(waiting_list.index(surgery), R.index(r), D.index(d), l), cat='Binary')

        for surgery in waiting_list:
            for surgeon in list_surgeon:
                num_int_var += 1
                q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d, l] = pulp.LpVariable("q({:},{:},{:},{:})".format(waiting_list.index(surgery), list_surgeon.index(surgeon), D.index(d), l), cat='Binary')

        for surgery1 in waiting_list:
            for surgery2 in waiting_list:
                if surgery1 != surgery2:
                    for r in R:
                        num_int_var += 1
                        y[surgery1.get_surgery_id(), surgery2.get_surgery_id(), r, d, l] = pulp.LpVariable("y({:},{:},{:},{:},{:})".format(waiting_list.index(surgery1), waiting_list.index(surgery2), R.index(r), D.index(d), l), cat='Binary')

        for surgery1 in waiting_list:
            for surgery2 in waiting_list:
                if surgery1 != surgery2:
                    for surgeon in list_surgeon:
                        num_int_var += 1
                        z[surgery1.get_surgery_id(), surgery2.get_surgery_id(), surgeon.get_surgeon_id(), d, l] = pulp.LpVariable("z({:},{:},{:},{:},{:})".format(waiting_list.index(surgery1), waiting_list.index(surgery2), list_surgeon.index(surgeon), D.index(d), l), cat='Binary')

        for surgeon in list_surgeon:
            num_int_var += 1
            n[surgeon.get_surgeon_id(), d, l] = pulp.LpVariable("n({:},{:})".format(list_surgeon.index(surgeon), D.index(d), l), cat='Binary')

        for surgery in waiting_list:
            num_con_var += 1
            ts[surgery.get_surgery_id(), l] = pulp.LpVariable("ts({:},{:})".format(waiting_list.index(surgery), l), lowBound=0, cat='Continuous')

        for r in R:
            num_con_var += 1
            msR[r, d, l] = pulp.LpVariable("msR({:},{:},{:})".format(R.index(r), D.index(d), l), lowBound=0, cat='Continuous')

        for r in R:
            num_con_var += 1
            ot[r, d, l] = pulp.LpVariable("ot({:},{:},{:})".format(R.index(r), D.index(d), l), lowBound=0, cat='Continuous')

        for surgeon in list_surgeon:
            tsS[surgeon.get_surgeon_id(), d, l] = pulp.LpVariable("tsS({:},{:},{:})".format(list_surgeon.index(surgeon), D.index(d), l), lowBound=0, cat='Continuous')

        for surgeon in list_surgeon:
            msS[surgeon.get_surgeon_id(), d, l] = pulp.LpVariable("msS({:},{:},{:})".format(list_surgeon.index(surgeon), D.index(d), l), lowBound=0, cat='Continuous')

        for surgeon in list_surgeon:
            num_con_var += 1
            wt[surgeon.get_surgeon_id(), d, l] = pulp.LpVariable("wt({:},{:},{:})".format(list_surgeon.index(surgeon), D.index(d), l), lowBound=0, cat='Continuous')

        for surgeon in list_surgeon:
            ost[surgeon.get_surgeon_id(), d, l] = pulp.LpVariable("ost({:},{:},{:})".format(list_surgeon.index(surgeon), D.index(d), l), lowBound=0, cat='Continuous')

        ns[d, l] = pulp.LpVariable("ns({:},{:})".format(d, l), lowBound=0, cat='Continuous')

        operating_room_planning = pulp.LpProblem("ORP" + str(l), pulp.LpMinimize)
        objective = pulp.lpSum(((d - A[surgery.get_surgery_id()]) + max(d - Due[surgery.get_surgery_id()], 0)) * x[surgery.get_surgery_id(), r, d, l] * U[surgery.get_surgery_id()] for surgery in waiting_list for r in R)
        objective += (pulp.lpSum(((Due[surgery.get_surgery_id()] - A[surgery.get_surgery_id()]) + (len(D) + 1 - Due[surgery.get_surgery_id()])) * U[surgery.get_surgery_id()] for surgery in waiting_list) - pulp.lpSum(((Due[surgery.get_surgery_id()] - A[surgery.get_surgery_id()]) + (len(D) + 1 - Due[surgery.get_surgery_id()])) * U[surgery.get_surgery_id()] * x[surgery.get_surgery_id(), r, d, l] for surgery in waiting_list for r in R))
        objective += 1000 * (pulp.lpSum(U[surgery.get_surgery_id()] for surgery in waiting_list) - pulp.lpSum(U[surgery.get_surgery_id()] * x[surgery.get_surgery_id(), r, d, l] for surgery in waiting_list for r in R))
        objective += pulp.lpSum(n[surgeon.get_surgeon_id(), d, l] for surgeon in list_surgeon) + pulp.lpSum(ot[r, d, l] for r in R) + pulp.lpSum(ost[surgeon.get_surgeon_id(), d, l] for surgeon in list_surgeon) + 10000 * (ns[d, l])
        objective += 100 * pulp.lpSum(1 / dict_room_group_prob[g, r] * x[surgery.get_surgery_id(), r, d, l] for g in G for surgery in list(set(waiting_list) & set(dict_surgery_group[g])) for r in dict_room_group[g])
        operating_room_planning += objective
        for surgery in waiting_list:
            operating_room_planning += pulp.lpSum(x[surgery.get_surgery_id(), r, d, l] for r in R) <= 1
        for surgery in waiting_list:
            operating_room_planning += pulp.lpSum(q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d, l] for surgeon in list_surgeon) <= 1

        for surgery in waiting_list:
            operating_room_planning += pulp.lpSum(x[surgery.get_surgery_id(), r, d, l] for r in R) == pulp.lpSum(q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d, l] for surgeon in list_surgeon)

        for surgeon in list_surgeon:
            operating_room_planning += pulp.lpSum(q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d, l] for surgery in waiting_list) <= len(waiting_list) * n[surgeon.get_surgeon_id(), d, l]

        for surgeon in list_surgeon:
            operating_room_planning += n[surgeon.get_surgeon_id(), d, l] <= pulp.lpSum(q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d, l] for surgery in waiting_list)

        for surgeon in list_surgeon:
            operating_room_planning += pulp.lpSum(q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d, l] for surgery in waiting_list) <= 3

        for r in R:
            operating_room_planning += pulp.lpSum((TP[surgery.get_surgery_id()] + TS[surgery.get_surgery_id()] + TC[surgery.get_surgery_id()]) * x[surgery.get_surgery_id(), r, d, l] for surgery in waiting_list) - T <= ot[r, d, l]
        """
        for g in G:
            for r in dict_room_group[g]:
                surgery_group = list(set(waiting_list) & set(dict_surgery_group[g]))
                if len(surgery_group) > 0:
                    operating_room_planning += 0.95 * dict_room_group_prob[g, r] <= pulp.lpSum(x[surgery.get_surgery_id(), r, d, l] for surgery in surgery_group) / len(surgery_group)
                    operating_room_planning += 1.05 * dict_room_group_prob[g, r] >= pulp.lpSum(x[surgery.get_surgery_id(), r, d, l] for surgery in surgery_group) / len(surgery_group)
        """
        for g in G:
            for surgery in list(set(dict_surgery_group[g]) & set(waiting_list)):
                for r in list(set(R) - set(dict_room_group[g])):
                    operating_room_planning += x[surgery.get_surgery_id(), r, d, l] == 0

        for g in G:
            for surgery in list(set(dict_surgery_group[g]) & set(waiting_list)):
                for surgeon in list(set(list_surgeon) - set(dict_surgeon_group[g])):
                    operating_room_planning += q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d, l] == 0

        for r in R:
            operating_room_planning += ot[r, d, l] <= O_max

        for surgeon in list_surgeon:
            operating_room_planning += ost[surgeon.get_surgeon_id(), d, l] >= pulp.lpSum(TS[surgery.get_surgery_id()] * q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d, l] for surgery in waiting_list) - 1 / len(list_surgeon) * pulp.lpSum(TS[surgery.get_surgery_id()] for sugery in waiting_list)

        for surgeon in list_surgeon:
            operating_room_planning += -ost[surgeon.get_surgeon_id(), d, l] <= pulp.lpSum(TS[surgery.get_surgery_id()] * q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d, l] for surgery in waiting_list) - 1 / len(list_surgeon) * pulp.lpSum(TS[surgery.get_surgery_id()] for sugery in waiting_list)

        operating_room_planning += ns[d, l] >= pulp.lpSum(x[surgery.get_surgery_id(), r, d, l] for surgery in waiting_list for r in R) - int(len(list_surgery) / len(D))
        operating_room_planning += -ns[d, l] <= pulp.lpSum(x[surgery.get_surgery_id(), r, d, l] for surgery in waiting_list for r in R) - int(len(list_surgery) / len(D))
        if l > 1:
            for i in range(1, l):
                for surgery1 in waiting_list:
                    for surgery2 in waiting_list:
                        if surgery1 != surgery2:
                            operating_room_planning += pulp.lpSum(x[surgery1.get_surgery_id(), r, d, l] for r in R if x[surgery1.get_surgery_id(), r, d, i].value() == 1) + pulp.lpSum(x[surgery2.get_surgery_id(), r, d, l] for r in R if x[surgery2.get_surgery_id(), r, d, i].value() == 1) + pulp.lpSum(q[surgery1.get_surgery_id(), surgeon.get_surgeon_id(), d, l] for surgeon in list_surgeon if q[surgery1.get_surgery_id(), surgeon.get_surgeon_id(), d, i].value() == 1) + pulp.lpSum(q[surgery2.get_surgery_id(), surgeon.get_surgeon_id(), d, l] for surgeon in list_surgeon if q[surgery2.get_surgery_id(), surgeon.get_surgeon_id(), d, i].value() == 1) <= 3

        result_status = operating_room_planning.solve(solver)
        print(pulp.LpStatus[result_status])
        operating_room_scheduling = pulp.LpProblem("ORS", pulp.LpMinimize)
        operating_room_scheduling += pulp.lpSum(ot[r, d, l] for r in R) + pulp.lpSum(wt[surgeon.get_surgeon_id(), d, l] for surgeon in list_surgeon) + pulp.lpSum(U[surgery.get_surgery_id()] * ts[surgery.get_surgery_id(), l] for surgery in waiting_list)

        for surgery1 in waiting_list:
            for surgery2 in waiting_list:
                if surgery1 != surgery2:
                    for r in R:
                        if x[surgery1.get_surgery_id(), r, d, l].value() == 1 and x[surgery2.get_surgery_id(), r, d, l].value() == 1:
                            operating_room_scheduling += y[surgery1.get_surgery_id(), surgery2.get_surgery_id(), r, d, l] + y[surgery2.get_surgery_id(), surgery1.get_surgery_id(), r, d, l] == 1

        for surgery1 in waiting_list:
            for surgery2 in waiting_list:
                if surgery1 != surgery2:
                    for surgeon in list_surgeon:
                        if q[surgery1.get_surgery_id(), surgeon.get_surgeon_id(), d, l].value() == 1 and q[surgery2.get_surgery_id(), surgeon.get_surgeon_id(), d, l].value() == 1:
                            operating_room_scheduling += z[surgery1.get_surgery_id(), surgery2.get_surgery_id(), surgeon.get_surgeon_id(), d, l] + z[surgery2.get_surgery_id(), surgery1.get_surgery_id(), surgeon.get_surgeon_id(), d, l] == 1

        for surgery in waiting_list:
            operating_room_scheduling += ts[surgery.get_surgery_id(), l] >= TP[surgery.get_surgery_id()]

        for surgery1 in waiting_list:
            for surgery2 in waiting_list:
                if surgery1 != surgery2:
                    for r in R:
                        if x[surgery1.get_surgery_id(), r, d, l].value() == 1 and x[surgery2.get_surgery_id(), r, d, l].value() == 1:
                            operating_room_scheduling += ts[surgery2.get_surgery_id(), l] - TP[surgery2.get_surgery_id()] >= ts[surgery1.get_surgery_id(), l] + TS[surgery1.get_surgery_id()] + TC[surgery1.get_surgery_id()] - M * (1 - y[surgery1.get_surgery_id(), surgery2.get_surgery_id(), r, d, l])

        for surgery1 in waiting_list:
            for surgery2 in waiting_list:
                if surgery1 != surgery2:
                    for surgeon in list_surgeon:
                        if q[surgery1.get_surgery_id(), surgeon.get_surgeon_id(), d, l].value() == 1 and q[surgery2.get_surgery_id(), surgeon.get_surgeon_id(), d, l].value() == 1:
                            operating_room_scheduling += ts[surgery2.get_surgery_id(), l] >= ts[surgery1.get_surgery_id(), l] + TS[surgery1.get_surgery_id()] + PT - M * (1 - z[surgery1.get_surgery_id(), surgery2.get_surgery_id(), surgeon.get_surgeon_id(), d, l])

        for surgery in waiting_list:
            for surgeon in list_surgeon:
                if q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d, l].value() == 1:
                    operating_room_scheduling += tsS[surgeon.get_surgeon_id(), d, l] <= ts[surgery.get_surgery_id(), l]

        for surgery in waiting_list:
            for r in R:
                if x[surgery.get_surgery_id(), r, d, l].value() == 1:
                    operating_room_scheduling += msR[r, d, l] >= ts[surgery.get_surgery_id(), l] + TS[surgery.get_surgery_id()] + TC[surgery.get_surgery_id()]

        for surgery in waiting_list:
            for surgeon in list_surgeon:
                if q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d, l].value() == 1:
                    operating_room_scheduling += msS[surgeon.get_surgeon_id(), d, l] >= ts[surgery.get_surgery_id(), l] + TS[surgery.get_surgery_id()]

        for r in R:
            operating_room_scheduling += ot[r, d, l] >= msR[r, d, l] - T

        for r in R:
            operating_room_scheduling += ot[r, d, l] <= O_max

        for surgeon in list_surgeon:
            operating_room_scheduling += wt[surgeon.get_surgeon_id(), d, l] >= msS[surgeon.get_surgeon_id(), d, l] - tsS[surgeon.get_surgeon_id(), d, l] - pulp.lpSum(TS[surgery.get_surgery_id()] * q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d, l].value() for surgery in waiting_list)

        result_status2 = operating_room_scheduling.solve(solver)
        if pulp.LpStatus[result_status2] == "Infeasible":
            print(pulp.LpStatus[result_status2])
            continue
        else:
            feasiblity_criteria = True
            L.append(l)
            for surgery in waiting_list:
                for r in R:
                    if x[surgery.get_surgery_id(), r, d, l].value() == 1:
                        x_val[surgery.get_surgery_id(), r, d] = 1
                        ts_val[surgery.get_surgery_id()] = ts[surgery.get_surgery_id(), l].value()

            for r in R:
                ot_val[r, d] = ot[r, d, l].value()
            for surgeon in list_surgeon:
                wt_val[surgeon.get_surgeon_id(), d] = wt[surgeon.get_surgeon_id(), d, l].value()
                if n[surgeon.get_surgeon_id(), d, l].value() == 1:
                    n_val[surgeon.get_surgeon_id(), d] = 1

    """
    print("式")
    print("--------")
    print(operating_room_scheduling)
    print("--------")
    print("")

    print("計算結果")
    print("********")

    print("最適性 = {:}, 目的関数値 = {:}"
          .format(pulp.LpStatus[result_status], pulp.value(operating_room_scheduling.objective)))
    """

    scheduled_list = []
    scheduled_surgeon_list = []
    for surgery in waiting_list:
        for r in R:
            if x[surgery.get_surgery_id(), r, d, l].value() == 1:
                scheduled_list.append(surgery)
        for surgeon in list_surgeon:
            if q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d, l].value() == 1:
                scheduled_surgeon_list.append(surgeon)
    scheduled_surgeon_list = list(set(scheduled_surgeon_list))
    sheet = book.worksheets[d - 1]
    sheet2 = book2.worksheets[d - 1]
    sheet.title = 'surgery_schedule_h1' + str(d)
    sheet2.title = 'surgeon_schedule_h1' + str(d)
    if D.index(d) < len(D) - 1:
        book.create_sheet()
        book2.create_sheet()
    create_excel_file(sheet, scheduled_list, R, d, l)
    create_excel_file2(sheet2, scheduled_list, scheduled_surgeon_list, d, l)
    with open('solution' + str(d) + '.csv', 'w') as csv_file:
        fieldnames = ['surgery id', 'surgeon id', 'room', 'group', 'date', 'release date', 'due date', 'preparation time', 'surgery time', 'cleaning time', 'entry time', 'start time', 'end time', 'exit time']
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        for r in R:
            list_surgery_room = classify_room_date(waiting_list, r, d, l)
            for surgery in list_surgery_room:
                for surgeon in list_surgeon:
                    if q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d, l].value() == 1 and x[surgery.get_surgery_id(), r, d, l].value() == 1:
                        start_int = ts[surgery.get_surgery_id(), l].value()
                        entry_int = start_int - TP[surgery.get_surgery_id()]
                        end_int = start_int + TS[surgery.get_surgery_id()]
                        exit_int = end_int + TC[surgery.get_surgery_id()]
                        entry_time = to_time_of_day(entry_int)
                        start_time = to_time_of_day(start_int)
                        end_time = to_time_of_day(end_int)
                        exit_time = to_time_of_day(exit_int)
                        writer.writerow({'surgery id': surgery.get_surgery_id(), 'surgeon id': surgeon.get_surgeon_id(), 'room': r, 'group': surgery.get_group(), 'date': d, 'release date': surgery.get_release_date(), 'due date': surgery.get_due_date(), 'preparation time': TP[surgery.get_surgery_id()], 'surgery time': TS[surgery.get_surgery_id()], 'cleaning time': TC[surgery.get_surgery_id()], 'entry time': entry_time, 'start time': start_time, 'end time': end_time, 'exit time': exit_time})
    waiting_list = list(set(waiting_list) - set(scheduled_list))
end = time.time()
print("残り手術数={:}".format(len(waiting_list) + len(list_surgery_copied)))
print("計算時間 = {:}".format(end - start))
print("手術数 = {:}".format(len(list_surgery)))
print("外科医の人数 = {:}".format(len(list_surgeon)))
def objective_function(lst_surgery, lst_surgeon, lst_room, lst_date):
    objective = pulp.lpSum(((d - A[surgery.get_surgery_id()]) + max(d - Due[surgery.get_surgery_id()], 0)) * x_val[surgery.get_surgery_id(), r, d] * U[surgery.get_surgery_id()] for surgery in lst_surgery for r in lst_room for d in lst_date)
    objective += pulp.lpSum(((Due[surgery.get_surgery_id()] - A[surgery.get_surgery_id()]) + (len(lst_date) + 1 - Due[surgery.get_surgery_id()])) * U[surgery.get_surgery_id()] for surgery in lst_surgery) - pulp.lpSum(((Due[surgery.get_surgery_id()] - A[surgery.get_surgery_id()]) + (len(lst_date) + 1 - Due[surgery.get_surgery_id()])) * U[surgery.get_surgery_id()] * x_val[surgery.get_surgery_id(), r, d] for surgery in lst_surgery for r in lst_room for d in lst_date)
    objective += 1000 * (pulp.lpSum(U[surgery.get_surgery_id()] for surgery in lst_surgery) - pulp.lpSum(U[surgery.get_surgery_id()] * x_val[surgery.get_surgery_id(), r, d] for surgery in lst_surgery for r in R for d in lst_date))
    objective += pulp.lpSum(n_val[surgeon.get_surgeon_id(), d] for surgeon in lst_surgeon for d in lst_date) + pulp.lpSum(ot_val[r, d] for r in lst_room for d in lst_date)
    objective += pulp.lpSum(wt_val[surgeon.get_surgeon_id(), d] for surgeon in lst_surgeon for d in lst_date) + pulp.lpSum(U[surgery.get_surgery_id()] * ts_val[surgery.get_surgery_id()] for surgery in lst_surgery)
    objective += pulp.lpSum(1 / dict_room_group_prob[g, r] * x_val[surgery.get_surgery_id(), r, d] for g in G for surgery in list(set(list_surgery) & set(dict_surgery_group[g])) for r in dict_room_group[g])
    return objective

objective_value = objective_function(list_surgery, list_surgeon, R, D)
print("目的関数値={:}".format(objective_value))
over_time = sum(ot_val[r, d] for r in R for d in D)
print("残業時間={:}".format(over_time))
book.save(url)
book.close()
book2.save(url2)
book2.close()
graph.create_ganttchart(url)
graph.create_surgeon_gantt_chart(url2)
