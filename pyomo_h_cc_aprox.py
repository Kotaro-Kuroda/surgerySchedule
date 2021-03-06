import pyomo.environ as pyo
import os
import numpy as np
import time
import openpyxl as px
import graph
import copy
import random
import instance
from scipy.stats import norm

home = os.environ['HOME']
directory_path = home + '/Documents/data-july'
path = directory_path + '/operations_with_jisseki_remake.csv'
surgeon_info_path = directory_path + '/surgeon_info.csv'
file_path = home + '/Documents/surgerySchedule/room_schedule_pyomo_heuristic1.xlsx'
file_path2 = home + '/Documents/surgerySchedule/surgeon_schedule_pyomo_heuristic1.xlsx'
surgeries_path = directory_path + '/surgeries.csv'
distribution_path = home + '/Documents/surgerySchedule/distribution.csv'
num_surgery = 4
num_date = 1
T = 480
PT = 15
M = 10 ** 5
O_max = 120
alpha = 0.95
seed = 1
surgery_instance = instance.SurgeryInstance(path, surgeon_info_path, num_surgery, num_date, seed, surgeries_path, distribution_path)
list_room, list_surgery, list_surgeon, list_date, list_group = surgery_instance.get_sets()
distribution_dict = surgery_instance.get_dict_distribution()
waiting_list = []
list_surgery_copied = copy.copy(list_surgery)
dict_surgery_id_room = surgery_instance.get_room_surgery_dict()

m1 = 10
m2 = 30
m3 = 10
m4 = 30
m5 = 1000
m6 = 20
m7 = 5
m8 = 10
m9 = 100
m10 = 10

x_val = {}
q_val = {}
ts_val = {}
ot_val = {}
wt_val = {}

for surgery in list_surgery:
    ts_val[surgery] = 0
for surgery in list_surgery:
    for r in list_room:
        for d in list_date:
            x_val[surgery, r, d] = 0
for surgery in list_surgery:
    for surgeon in list_surgeon:
        for d in list_date:
            q_val[surgery, surgeon, d] = 0

for r in list_room:
    for d in list_date:
        ot_val[r, d] = 0

for surgeon in list_surgeon:
    for d in list_date:
        wt_val[surgeon, d] = 0
scheduled_surgery = []

dict_not_available_room = {}
for surgery in list_surgery:
    types = surgery.get_surgery_type()
    available_room = set(list_room)
    for s_type in types:
        if s_type in dict_surgery_id_room.keys():
            available_room = available_room & set(dict_surgery_id_room[s_type])
        else:
            available_room = available_room & set(list_room)
    not_available_room = set(list_room) - available_room
    dict_not_available_room[surgery] = list(not_available_room)


def aprox_norm_cdf(x):
    return 1 - 1 / np.sqrt(2 * np.pi) * pyo.exp(-x**2 / 2) / (0.226 + 0.64 * x + 0.33 * pyo.sqrt(x**2 + 3))


def sigma(model, x, U, D):
    return np.log((sum(D[s] * x[s] for s in model.S) + (sum(U[s] * x[s] for s in model.S)) ** 2) / (sum(U[s] * x[s] for s in model.S) ** 2))
    # return np.log((np.dot(D, x) + np.dot(U, x) ** 2) / (np.dot(U, x) ** 2))
def log_normal_cc(model, x, U, D, ot):
    if set(x.values()) == {0}:
        return -np.inf
    return np.log(sum(U[s] * x[s] for s in model.S)) - 1 / 2 * sigma(model, x, U, D) + norm.ppf(alpha) * np.sqrt(sigma(model, x, U, D)) - np.log(T + ot)
    # return np.log(np.dot(U, x)) - 1/2 * sigma(x, U, D) + norm.ppf(0.95) * np.sqrt(sigma(x, U, D)) - np.log(T + ot)

def grad(model, x, U, D, ot):
    grad_dict = {}
    if set(x.values()) == {0}:
        for s in model.S:
            grad_dict[s] = -np.inf
    else:
        for s1 in model.S:
            grad_dict[s1] = 2 * U[s1] / (sum(U[s] * x[s] for s in model.S)) - 1 / 2 * (D[s1] + 2 * U[s1] * (sum(U[s] * x[s] for s in model.S))) / (sum(D[s] * x[s] for s in model.S) + (sum(U[s] * x[s] for s in model.S)) ** 2) + norm.ppf(alpha) / 2 * sigma(model, x, U, D) ** (-0.5) * (sum(U[s] * x[s] for s in model.S) * D[s1] - 2 * sum(D[s] * x[s] for s in model.S) * U[s1]) / ((sum(D[s] * x[s] for s in model.S) + sum(U[s] * x[s] for s in model.S) ** 2) * sum(U[s] * x[s] for s in model.S))

    grad_dict['ot'] = -1 / np.log(ot + T)
    return grad_dict
    # return np.append(2 * U / (np.dot(U, x)) - 1/2 * (D + 2 * U * np.dot(U, x)) / (np.dot(D, x) + np.dot(U, x) ** 2) + norm.ppf(0.95) / 2 * sigma(x, U, D) ** (-0.5) * ((np.dot(U, x) * D - 2 * np.dot(D, x) * U) / ((np.dot(D, x) + np.dot(U, x) ** 2) * np.dot(U, x))), - 1/(T + ot))

def get_right_hand(alpha, u, d):
    if u > 0 and d > 0:
        sigma = np.log(d / (u ** 2) + 1)
        mu = np.log(u) - sigma / 2
        return np.exp(np.sqrt(sigma) * norm.ppf(alpha) + mu)
    else:
        return 0
def get_left_hand(alpha, u, d):
    return pyo.log(u) - 1 / 2 * pyo.log(d + u ** 2) + 1 / 2 * pyo.log(u ** 2) + pyo.sqrt(pyo.log(d + u ** 2) - pyo.log(u ** 2)) * norm.ppf(alpha)

def combS_init(model):
    return ((s1, s2) for s1 in model.S for s2 in model.S if s1 != s2)

def R_init(model, s):
    return dict_not_available_room[s]

def rule1(model, s):
    return sum(model.x[s, r] for r in model.R) <= 1

def rule2(model, s):
    return sum(model.q[s, k] for k in model.K) <= 1

def rule3(model, s):
    return sum(model.x[s, r] for r in model.R) == sum(model.q[s, k] for k in model.K)

"""prod = 1
                for s in model.S:
                    prod *= (1 - model.x[s, r])"""
def rule4(model, r):
    # return sum((s.get_preparation_time() + s.get_surgery_time() + s.get_cleaning_time()) * model.x[s, r] for s in model.S) <= model.ot[r] + T
    a = {}
    if l == 1:
        for s in model.S:
            a[s] = random.randint(0, 1)
            ot1 = random.randint(0, 120)
    else:
        for s in model.S:
            a[s] = x_val[s, r, d, l - 1]
            ot1 = ot_val[r, d, l - 1]
    U = {}
    D = {}
    for s in model.S:
        U[s] = distribution_dict[s.get_group(), 'total_mean']
        D[s] = distribution_dict[s.get_group(), 'total_variance']
    if set(a.values()) == {0}:
        min_mean = min(list(U.values()))
        index = [k for k, v in U.items() if v == min_mean][0]
        a[index] = 1
    grad_dict = grad(model, a, U, D, ot1)
    aprox = log_normal_cc(model, a, U, D, ot1) + sum(grad_dict[s] * (model.x[s, r] - a[s]) for s in model.S) + grad_dict['ot'] * (model.ot[r] - ot1)
    return aprox <= 0
def rule5(model, s):
    set_not_available_room = list(model.R_not_available[s])
    if len(set_not_available_room) > 0:
        return sum(model.x[s, r] for r in set_not_available_room) == 0
    else:
        return pyo.Constraint.NoConstraint

def rule6(model, r):
    return model.ot[r] <= O_max

def rule9(model, k):
    return model.ost[k] >= sum(model.q[s, k] for s in model.S) - len(model.S) / len(model.K)

def rule10(model, k):
    return -model.ost[k] <= sum(model.q[s, k] for s in model.S) - len(model.S) / len(model.K)

def rule7(model, r):
    return model.ar[r] >= sum(model.x[s, r] for s in model.S) - len(model.S) / len(model.R)

def rule8(model, r):
    return -model.ar[r] <= sum(model.x[s, r] for s in model.S) - len(model.S) / len(model.R)


def rule11(model):
    return model.ns >= sum(model.x[s, r] for s in model.S for r in model.R) - len(list_surgery) / len(list_date)

def rule12(model):
    return -model.ns <= sum(model.x[s, r] for s in model.S for r in model.R) - len(list_surgery) / len(list_date)

def fc1(model, s1, s2, j):
    return sum(model.x[s1, r] for r in model.R if x_val[s1, r, d, j] == 1) + sum(model.x[s2, r] for r in model.R if x_val[s2, r, d, j] == 1) + sum(model.q[s1, k] for k in model.K if q_val[s1, k, d, j] == 1) + sum(model.q[s2, k] for k in model.K if q_val[s2, k, d, j] == 1) <= 3


def objective(model):
    ob = sum(B[s, r] * ((m1 * (d - s.get_release_date()) + m2 * max(d - s.get_due_date(), 0)) * model.x[s, r] * s.get_priority()) for s in model.S for r in model.R)
    ob += sum(m3 * ((s.get_due_date() - s.get_release_date()) + m4 * (len(list_date) + 1 - s.get_due_date())) * (1 - sum(model.x[s, r] for r in model.R)) * s.get_priority() for s in model.S)
    ob += m5 * (sum(s.get_priority() * (1 - sum(model.x[s, r] for r in model.R)) for s in model.S))
    ob += m6 * sum(model.A[r] * model.ot[r] for r in model.R)
    ob += m9 * sum(model.ost[k] for k in model.K) + m10 * model.ns
    ob += 100 * sum(model.ar[r] for r in model.R)
    return ob

def rule_s1(model, s1, s2, r):
    if x_val[s1, r, d, l] == 1 and x_val[s2, r, d, l] == 1:
        return model.y[s1, s2, r] + model.y[s2, s1, r] == 1
    else:
        return pyo.Constraint.NoConstraint

def rule_s2(model, s1, s2, k):
    if q_val[s1, k, d, l] == 1 and q_val[s2, k, d, l] == 1:
        return model.z[s1, s2, k] + model.z[s2, s1, k] == 1
    else:
        return pyo.Constraint.NoConstraint

def rule_s3_2(model):
    left_hand = sum(aprox_norm_cdf((pyo.log(model.ts[s2] - model.ts[s1] + M * (1 - model.y[s1, s2, r])) - (distribution_dict[s2.get_group(), "preparation_mean"] + distribution_dict[s1.get_group(), "surgery_mean"] + distribution_dict[s1.get_group(), "cleaning_mean"])) / (pyo.sqrt(distribution_dict[s2.get_group(), "preparation_variance"] + distribution_dict[s1.get_group(), "surgery_variance"] + distribution_dict[s1.get_group(), "cleaning_variance"]))) for s1 in model.S for s2 in model.S for r in model.R if s1 != s2)
    return left_hand >= np.log(alpha)
def rule_s3(model, s1, s2, r):
    # return sum(aprox_norm_cdf((pyo.log(model.ts[s2] - model.ts[s1] + M * (1 - model.y[s1, s2, r])) - (distribution_dict[s2.get_group(), "preparation_mean"] + distribution_dict[s1.get_group(), "surgery_mean"] + distribution_dict[s1.get_group(), "cleaning_mean"]))) / pyo.sqrt(distribution_dict[s2.get_group(), "preparation_variance"] + distribution_dict[s1.get_group(), "surgery_variance"] + distribution_dict[s1.get_group(), "cleaning_variance"]) for s1 in model.S for s2 in model.S if s1 != s2 for r in model.R) >= np.log(alpha)
    g1 = s1.get_group()
    g2 = s2.get_group()
    mean = distribution_dict[g2, 'preparation_mean'] + distribution_dict[g1, 'surgery_mean'] + distribution_dict[g1, 'cleaning_mean']
    variance = distribution_dict[g2, 'preparation_variance'] + distribution_dict[g1, 'surgery_variance'] + distribution_dict[g1, 'cleaning_variance']
    return model.ts[s2] - model.ts[s1] + M * (1 - model.y[s1, s2, r]) >= get_right_hand(alpha, mean, variance)
def rule_s4(model, s1, s2, k):
    g = s1.get_group()
    mean = distribution_dict[g, 'surgery_mean']
    variance = distribution_dict[g, 'surgery_variance']
    return model.ts[s2] - model.ts[s1] - PT + M * (1 - model.z[s1, s2, k]) >= get_right_hand(alpha, mean, variance)

def rule_s5(model, s):
    g = s.get_group()
    mean = distribution_dict[g, 'preparation_mean']
    variance = distribution_dict[g, 'preparation_variance']
    if sum(x_val[s, r, d, l] for r in model.R) == 1:
        return model.ts[s] >= get_right_hand(alpha, mean, variance)
    else:
        return pyo.Constraint.NoConstraint
def rule_s6(model, s, k):
    if q_val[s, k, d, l] == 1:
        return model.tsS[k] <= model.ts[s] + M * (1 - q_val[s, k, d, l])
    else:
        return pyo.Constraint.NoConstraint

def rule_s7(model, s, r):
    g = s.get_group()
    mean = distribution_dict[g, 'surgery_mean'] + distribution_dict[g, 'cleaning_mean']
    variance = distribution_dict[g, 'surgery_variance'] + distribution_dict[g, 'cleaning_variance']
    if x_val[s, r, d, l] == 1:
        return model.msR[r] - model.ts[s] >= get_right_hand(alpha, mean, variance)
    else:
        return pyo.Constraint.NoConstraint

def rule_s8(model, s, k):
    g = s.get_group()
    mean = distribution_dict[g, 'surgery_mean']
    variance = distribution_dict[g, 'surgery_variance']
    if q_val[s, k, d, l] == 1:
        return model.msS[k] - model.ts[s] >= get_right_hand(alpha, mean, variance)
    else:
        return pyo.Constraint.NoConstraint

def rule_s9(model, r):
    return model.ot[r] >= model.msR[r] - T

def rule_s10(model, r):
    return model.ot[r] <= O_max

def rule_s11(model, k):
    mean = sum(distribution_dict[s.get_group(), 'surgery_mean'] * q_val[s, k, d, l] for s in model.S)
    variance = sum(distribution_dict[s.get_group(), 'surgery_variance'] * q_val[s, k, d, l] ** 2 for s in model.S)
    return model.msS[k] - model.tsS[k] - model.wt[k] <= get_right_hand(1 - alpha, mean, variance)

def objective_s(model):
    return m6 * sum(model.ot[r] for r in model.R) + m7 * sum(model.wt[k] for k in model.K) + m8 * sum(s.get_priority() * model.ts[s] for s in model.S)
opt = pyo.SolverFactory("cplex")

# opt.options["limits/gap"] = 0.01

start = time.time()
for d in list_date:
    feasiblity_criteria = False
    print("date = " + str(d))
    for surgery in list_surgery_copied[:]:
        if surgery.get_release_date() <= d and surgery not in waiting_list:
            waiting_list.append(surgery)
            list_surgery_copied.remove(surgery)

    l = 0
    if len(waiting_list) == 0:
        continue
    while not feasiblity_criteria:
        l += 1
        print("loop = " + str(l))
        model = pyo.ConcreteModel()
        model.S = pyo.Set(initialize=waiting_list)
        model.R = pyo.Set(initialize=list_room)
        model.R_not_available = pyo.Set(model.S, initialize=R_init)
        model.K = pyo.Set(initialize=list_surgeon)
        model.comb_S = pyo.Set(initialize=combS_init)
        model.x = pyo.Var(model.S, model.R, domain=pyo.Binary)
        model.q = pyo.Var(model.S, model.K, domain=pyo.Binary)
        # model.p = pyo.Var(model.R, domain=pyo.Binary)
        model.ot = pyo.Var(model.R, domain=pyo.NonNegativeReals)
        model.ost = pyo.Var(model.K, domain=pyo.NonNegativeReals)
        model.ns = pyo.Var(domain=pyo.NonNegativeReals)
        model.ar = pyo.Var(model.R, domain=pyo.NonNegativeReals)

        A = {}
        for r in model.R:
            A[r] = 1 + 0.05 * np.random.random()
        model.A = pyo.Param(model.R, initialize=A)
        B = {}
        for s in model.S:
            for r in model.R:
                B[s, r] = 1 + 0.05 * np.random.random()
        model.B = pyo.Param(model.S, model.R, initialize=B)
        model.constrain1 = pyo.Constraint(model.S, rule=rule1)
        model.constraint2 = pyo.Constraint(model.S, rule=rule2)
        model.constraint3 = pyo.Constraint(model.S, rule=rule3)
        model.constraint5 = pyo.Constraint(model.S, rule=rule5)
        model.constraint4 = pyo.Constraint(model.R, rule=rule4)
        model.constraint6 = pyo.Constraint(model.R, rule=rule6)
        model.constraint9 = pyo.Constraint(model.K, rule=rule9)
        model.constraint10 = pyo.Constraint(model.K, rule=rule10)
        model.constraint11 = pyo.Constraint(rule=rule11)
        model.constraint12 = pyo.Constraint(rule=rule12)
        model.Constraint7 = pyo.Constraint(model.R, rule=rule7)
        model.Constraint8 = pyo.Constraint(model.R, rule=rule8)
        model.objective = pyo.Objective(rule=objective)
        if l == 1:
            res = opt.solve(model)
            if res.solver.termination_condition == 'infeasible':
                print("ORP:実行不可能")
                l -= 1
                continue
            for r in model.R:
                ot_val[r, d, l] = model.ot[r]()
            for s in model.S:
                for r in model.R:
                    x_val[s, r, d, l] = round(model.x[s, r]())
                    if x_val[s, r, d, l] == 1:
                        print('r={:}'.format(r))
                    # print(f"x[{s}, {r}] = {model.x[s, r]()}")

            for s in model.S:
                for k in model.K:
                    q_val[s, k, d, l] = round(model.q[s, k]())
                    if q_val[s, k, d, l] == 1:
                        print('k={:}'.format(k.get_surgeon_id()))
                    # print(f"q[{s}, {k}] = {model.q[s, k]()}")
        else:
            model.fc1 = pyo.Constraint(model.comb_S, pyo.RangeSet(l - 1), rule=fc1)
            res = opt.solve(model)
            print(res.solver.termination_condition)
            if res.solver.termination_condition == 'infeasible':
                print("ORP:実行不可能")
                continue
            for r in model.R:
                ot_val[r, d, l] = model.ot[r]()

            for s in model.S:
                for r in model.R:
                    x_val[s, r, d, l] = round(model.x[s, r]())
                    if x_val[s, r, d, l] == 1:
                        print('r={:}'.format(r))

            for s in model.S:
                for k in model.K:
                    q_val[s, k, d, l] = round(model.q[s, k]())
                    if q_val[s, k, d, l] == 1:
                        print('k={:}'.format(k.get_surgeon_id()))

        temp_scheduled_surgery = [s for s in model.S if sum(x_val[s, r, d, l] for r in model.R) == 1]
        temp_scheduled_room = [r for r in model.R if sum(x_val[s, r, d, l] for s in model.S) >= 1]
        temp_scheduled_surgeon = [k for k in model.K if sum(q_val[s, k, d, l] for s in model.S) >= 1]
        operating_room_scheduling = pyo.ConcreteModel()
        operating_room_scheduling.S = pyo.Set(initialize=waiting_list)
        operating_room_scheduling.R = pyo.Set(initialize=list_room)
        operating_room_scheduling.K = pyo.Set(initialize=list_surgeon)
        operating_room_scheduling.comb_S = pyo.Set(initialize=combS_init)
        operating_room_scheduling.y = pyo.Var(operating_room_scheduling.comb_S, operating_room_scheduling.R, domain=pyo.Binary)
        operating_room_scheduling.z = pyo.Var(operating_room_scheduling.comb_S, operating_room_scheduling.K, domain=pyo.Binary)
        operating_room_scheduling.ot = pyo.Var(operating_room_scheduling.R, domain=pyo.NonNegativeReals)
        operating_room_scheduling.ts = pyo.Var(operating_room_scheduling.S, domain=pyo.NonNegativeReals)
        operating_room_scheduling.msR = pyo.Var(operating_room_scheduling.R, domain=pyo.NonNegativeReals)
        operating_room_scheduling.tsS = pyo.Var(operating_room_scheduling.K, domain=pyo.NonNegativeReals)
        operating_room_scheduling.msS = pyo.Var(operating_room_scheduling.K, domain=pyo.NonNegativeReals)
        operating_room_scheduling.wt = pyo.Var(operating_room_scheduling.K, domain=pyo.NonNegativeReals)

        operating_room_scheduling.constraint1 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.R, rule=rule_s1)
        # operating_room_scheduling.constraint1_1 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.R, rule=rule_s1_2)
        operating_room_scheduling.constraint2 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.K, rule=rule_s2)
        # operating_room_scheduling.constraint2_2 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.K, rule=rule_s2_2)
        # operating_room_scheduling.constraint3 = pyo.Constraint(rule=rule_s3_2)
        operating_room_scheduling.constraint3 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.R, rule=rule_s3)
        operating_room_scheduling.constraint4 = pyo.Constraint(operating_room_scheduling.comb_S, operating_room_scheduling.K, rule=rule_s4)
        operating_room_scheduling.constraint5 = pyo.Constraint(operating_room_scheduling.S, rule=rule_s5)
        operating_room_scheduling.constraint6 = pyo.Constraint(operating_room_scheduling.S, operating_room_scheduling.K, rule=rule_s6)
        operating_room_scheduling.constraint7 = pyo.Constraint(operating_room_scheduling.S, operating_room_scheduling.R, rule=rule_s7)
        operating_room_scheduling.constraint8 = pyo.Constraint(operating_room_scheduling.S, operating_room_scheduling.K, rule=rule_s8)
        operating_room_scheduling.constraint9 = pyo.Constraint(operating_room_scheduling.R, rule=rule_s9)
        operating_room_scheduling.constraint10 = pyo.Constraint(operating_room_scheduling.R, rule=rule_s10)
        operating_room_scheduling.constraint11 = pyo.Constraint(operating_room_scheduling.K, rule=rule_s11)
        operating_room_scheduling.objective = pyo.Objective(rule=objective_s)
        res = pyo.SolverFactory("cplex").solve(operating_room_scheduling)
        print("ORS:{:}".format(res.solver.termination_condition))
        if res.solver.termination_condition == 'infeasible':
            continue
        else:
            feasiblity_criteria = True
            for s in model.S:
                for r in model.R:
                    x_val[s, r, d] = x_val[s, r, d, l]
                ts_val[s] = operating_room_scheduling.ts[s]()
                for k in model.K:
                    q_val[s, k, d] = q_val[s, k, d, l]

            for k in model.K:
                wt_val[k, d] = operating_room_scheduling.wt[k]()
            for r in model.R:
                ot_val[r, d] = operating_room_scheduling.ot[r]()
            for s in model.S:
                for r in model.R:
                    if x_val[s, r, d] == 1:
                        waiting_list.remove(s)
                        scheduled_surgery.append(s)
end = time.time()

def objective_function(lst_surgery, lst_surgeon, lst_room, lst_date):
    objective = sum((m1 * (d - surgery.get_release_date()) + m2 * max(d - surgery.get_due_date(), 0)) * x_val[surgery, r, d] * surgery.get_priority() for surgery in lst_surgery for r in lst_room for d in lst_date)
    objective += sum((m3 * (surgery.get_due_date() - surgery.get_release_date()) + m4 * (len(lst_date) + 1 - surgery.get_due_date())) * surgery.get_priority() * (1 - sum(x_val[surgery, r, d] for r in lst_room for d in lst_date)) for surgery in lst_surgery)
    objective += m5 * sum(surgery.get_priority() * (1 - sum(x_val[surgery, r, d] for r in lst_room for d in lst_date)) for surgery in lst_surgery)
    objective += m6 * sum(ot_val[r, d] for r in lst_room for d in lst_date)
    objective += m7 * sum(wt_val[surgeon, d] for surgeon in lst_surgeon for d in lst_date) + m8 * sum(surgery.get_priority() * ts_val[surgery] for surgery in lst_surgery)
    return objective
print("計算時間={:}".format(end - start))
print("残業時間={:}".format(sum(ot_val[r, d] for r in list_room for d in list_date)))
print('スケジュールされた手術{:}'.format(len(scheduled_surgery)))
print(objective_function(list_surgery, list_surgeon, list_room, list_date))
print(len(list_surgery), len(list_surgeon), len(list_room), len(list_date))
def sort(list_surgery, ts_val):
    if len(list_surgery) <= 1:
        return list_surgery
    pivot = np.random.choice(list_surgery)
    pivot_val = ts_val[pivot]
    left = []
    right = []
    for i in range(len(list_surgery)):
        surgery = list_surgery[i]
        start_time = ts_val[surgery]
        if start_time < pivot_val:
            left.append(surgery)
        else:
            right.append(surgery)
    return sort(left, ts_val) + sort(right, ts_val)


def get_surgeries_in_r(list_surgery, room, date, x_val, ts_val):
    lst = []
    for surgery in list_surgery:
        if (surgery, room, date) in x_val.keys() and x_val[surgery, room, date] == 1:
            lst.append(surgery)
    return sort(lst, ts_val)


def get_surgeries_by_k(list_surgery, surgeon, date, q_val, ts_val):
    lst = []
    for surgery in list_surgery:
        if (surgery, surgeon, date) in q_val.keys() and q_val[surgery, surgeon, date] == 1:
            lst.append(surgery)

    return sort(lst, ts_val)


def create_excel_file(list_surgery, list_date, list_room, ts_val, x_val):
    book = px.Workbook()
    for l in range(len(list_date)):
        d = list_date[l]
        sheet = book.worksheets[l]
        sheet.title = 'surgery_schedule' + str(d)
        length = 0
        scheduled_surgery_d = []
        for s in list_surgery:
            for r in list_room:
                if (s, r, d) in x_val.keys() and x_val[s, r, d] == 1:
                    scheduled_surgery_d.append(s)
        for i in range(len(list_room)):
            for j in range(len(scheduled_surgery_d) * 4):
                r = list_room[i]
                sheet.cell(row=j + 2, column=list_room.index(r) + 2).value = 0
        for r in list_room:
            list_surgery_in_r = get_surgeries_in_r(scheduled_surgery_d, r, d, x_val, ts_val)
            sheet.cell(row=1, column=list_room.index(r) + 2).value = "手術室" + r
            for i in range(len(list_surgery_in_r) * 4):
                surgery = list_surgery_in_r[int(i / 4)]
                if i % 4 == 0:
                    sheet.cell(row=i + 2 + length, column=1).value = "空白"
                    if i == 0:
                        sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = ts_val[surgery] - surgery.get_random_preparation()
                    else:
                        prev_surgery = list_surgery_in_r[int(i / 4 - 1)]
                        entry_time = ts_val[surgery] - surgery.get_random_preparation()
                        prev_exit_time = ts_val[prev_surgery] + prev_surgery.get_random_surgery() + prev_surgery.get_random_cleaning()
                        sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = entry_time - prev_exit_time
                elif i % 4 == 1:
                    sheet.cell(row=i + 2 + length, column=1).value = "準備" + str(surgery)
                    sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = surgery.get_random_preparation()
                elif i % 4 == 2:
                    sheet.cell(row=i + 2 + length, column=1).value = "手術" + str(surgery)
                    sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = surgery.get_random_surgery()
                if i % 4 == 3:
                    sheet.cell(row=i + 2 + length, column=1).value = "清掃" + str(surgery)
                    sheet.cell(row=i + 2 + length, column=list_room.index(r) + 2).value = surgery.get_random_cleaning()
            length += len(list_surgery_in_r) * 4
        if l < len(list_date) - 1:
            book.create_sheet()
    book.save(file_path)
    book.close()

def create_excel_file2(list_surgery, list_date, list_surgeon, ts_val, q_val):
    book = px.Workbook()
    for l in range(len(list_date)):
        d = list_date[l]
        sheet = book.worksheets[l]
        sheet.title = 'surgery_schedule' + str(d)
        length = 0
        scheduled_surgery_d = []
        schedule_surgeon_d = []
        for surgery in list_surgery:
            for surgeon in list_surgeon:
                if q_val[surgery, surgeon, d] == 1:
                    scheduled_surgery_d.append(surgery)
                    if surgeon not in schedule_surgeon_d:
                        schedule_surgeon_d.append(surgeon)
        for i in range(len(schedule_surgeon_d)):
            for j in range(len(list_surgery) * 2):
                sheet.cell(row=j + 2, column=i + 2).value = 0
        for i in range(len(schedule_surgeon_d)):
            surgeon = schedule_surgeon_d[i]
            list_surgery_by_surgeon = get_surgeries_by_k(list_surgery, surgeon, d, q_val, ts_val)
            sheet.cell(row=1, column=i + 2).value = "外科医" + str(surgeon.get_surgeon_id())
            for j in range(len(list_surgery_by_surgeon) * 2):
                surgery = list_surgery_by_surgeon[int(j / 2)]
                if j % 2 == 0:
                    sheet.cell(row=j + 2 + length, column=1).value = "空白"
                    if j == 0:
                        sheet.cell(row=j + 2 + length, column=i + 2).value = ts_val[surgery]
                    else:
                        prev_surgery = list_surgery_by_surgeon[int(j / 2 - 1)]
                        sheet.cell(row=j + 2 + length, column=i + 2).value = ts_val[surgery] - (ts_val[prev_surgery] + prev_surgery.get_random_surgery() + prev_surgery.get_random_cleaning())
                else:
                    sheet.cell(row=j + 2 + length, column=1).value = "手術" + surgery.get_surgery_id()
                    sheet.cell(row=j + 2 + length, column=i + 2).value = surgery.get_random_surgery()
            length += len(list_surgery_by_surgeon) * 2
        if l < len(list_date) - 1:
            book.create_sheet()
    book.save(file_path2)
    book.close()

create_excel_file(scheduled_surgery, list_date, list_room, ts_val, x_val)
create_excel_file2(scheduled_surgery, list_date, list_surgeon, ts_val, q_val)
graph.create_ganttchart(file_path)
graph.create_surgeon_ganttchart(file_path2)
