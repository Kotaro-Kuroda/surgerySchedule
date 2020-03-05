import time
import pulp
import os
import openpyxl
import csv
import graph
import create_instance

operations = []
with open('/Users/kurodakotaro/Documents/data-july/operations.csv') as f:
    reader = csv.reader(f)
    operations = [row for row in reader]

home = os.environ['HOME']
url = home + '/Documents/solution/solution_orps.xlsx'
url2 = home + '/Documents/solution/solution2_orps.xlsx'

# 手術室の集合
R = list(set([operations[i][5] for i in range(1, len(operations))]))
G = list(set([operations[i][6] for i in range(1, len(operations))]))

num_int_var = 0
num_con_var = 0
num_constraint = 0
initial = time.time()
num_surgery = 100
num_date = 7
seed = 1
list_surgery, list_surgeon, D = create_instance.create_model(num_surgery, num_date, seed)
dict_surgery_group = {}
for g in G:
    list_surgery_group = []
    for surgery in list_surgery:
        if g == surgery.get_group():
            list_surgery_group.append(surgery)
    dict_surgery_group[g] = list_surgery_group

dict_surgeon_group = {}
for g in G:
    list_surgeon_group = []
    for surgeon in list_surgeon:
        if g == surgeon.get_group():
            list_surgeon_group.append(surgeon)
    dict_surgeon_group[g] = list_surgeon_group

dict_room_group = {}
for g in G:
    list_room_group = []
    for i in range(1, len(operations)):
        if g == operations[i][6]:
            list_room_group.append(operations[i][5])
    list_room_group = list(set(list_room_group))
    dict_room_group[g] = list_room_group

dict_room_group_prob = {}
for g in G:
    for r in dict_room_group[g]:
        dict_room_group_prob[g, r] = sum(1 for i in range(1, len(operations)) if operations[i][6] == g and operations[i][5] == r) / sum(1 for i in range(1, len(operations)) if operations[i][6] == g)

TS = {surgery.get_surgery_id(): surgery.get_surgery_time() for surgery in list_surgery}

TP = {surgery.get_surgery_id(): surgery.get_preparation_time() for surgery in list_surgery}

TC = {surgery.get_surgery_id(): surgery.get_cleaning_time() for surgery in list_surgery}

A = {surgery.get_surgery_id(): surgery.get_release_date() for surgery in list_surgery}

Due = {surgery.get_surgery_id(): surgery.get_due_date() for surgery in list_surgery}

U = {surgery.get_surgery_id(): surgery.get_priority() for surgery in list_surgery}

T = 480
PT = 15
M = 99999
ope_time = 9
x = {}
for surgery in list_surgery:
    for r in R:
        for d in D:
            num_int_var += 1
            x[surgery.get_surgery_id(), r, d] = pulp.LpVariable("x({:},{:},{:})".format(list_surgery.index(surgery), R.index(r), D.index(d)), cat='Binary')

q = {}
for surgery in list_surgery:
    for surgeon in list_surgeon:
        for d in D:
            num_int_var += 1
            q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d] = pulp.LpVariable("q({:},{:}, {:})".format(list_surgery.index(surgery), list_surgeon.index(surgeon), D.index(d)), cat='Binary')

y = {}
for surgery1 in list_surgery:
    for surgery2 in list_surgery:
        for r in R:
            for d in D:
                if surgery1 != surgery2:
                    num_int_var += 1
                    y[surgery1.get_surgery_id(), surgery2.get_surgery_id(), r, d] = pulp.LpVariable("y({:},{:},{:},{:})".format(list_surgery.index(surgery1), list_surgery.index(surgery2), R.index(r), D.index(d)), cat='Binary')
z = {}
for surgery1 in list_surgery:
    for surgery2 in list_surgery:
        for surgeon in list_surgeon:
            for d in D:
                if surgery1 != surgery2:
                    num_int_var += 1
                    z[surgery1.get_surgery_id(), surgery2.get_surgery_id(), surgeon.get_surgeon_id(), d] = pulp.LpVariable("z({:},{:},{:},{:})".format(list_surgery.index(surgery1), list_surgery.index(surgery2), list_surgeon.index(surgeon), D.index(d)), cat='Binary')

n = {}
for surgeon in list_surgeon:
    for d in D:
        num_int_var += 1
        n[surgeon.get_surgeon_id(), d] = pulp.LpVariable("n({:},{:})".format(list_surgeon.index(surgeon), D.index(d)), cat='Binary')
ts = {}
for surgery in list_surgery:
    for d in D:
        num_con_var += 1
        ts[surgery.get_surgery_id()] = pulp.LpVariable("ts({:})".format(list_surgery.index(surgery)), lowBound=0, cat='Continuous')


msR = {}
for r in R:
    for d in D:
        num_con_var += 1
        msR[r, d] = pulp.LpVariable("msR({:}, {:})".format(R.index(r), D.index(d)), lowBound=0, cat='Integer')

ot = {}
for r in R:
    for d in D:
        num_con_var += 1
        ot[r, d] = pulp.LpVariable("ot({:}, {:})".format(R.index(r), D.index(d)), lowBound=0, cat='Integer')

tsS = {}
for surgeon in list_surgeon:
    for d in D:
        tsS[surgeon.get_surgeon_id(), d] = pulp.LpVariable("tsS({:},{:})".format(list_surgeon.index(surgeon), D.index(d)), lowBound=0, cat='Integer')

msS = {}
for surgeon in list_surgeon:
    for d in D:
        msS[surgeon.get_surgeon_id(), d] = pulp.LpVariable("msS({:},{:})".format(list_surgeon.index(surgeon), D.index(d)), lowBound=0, cat='Integer')

wt = {}
for surgeon in list_surgeon:
    for d in D:
        num_con_var += 1
        wt[surgeon.get_surgeon_id(), d] = pulp.LpVariable("wt({:}, {:})".format(list_surgeon.index(surgeon), D.index(d)), lowBound=0, cat='Integer')

problem = pulp.LpProblem("ORSP", pulp.LpMinimize)

def objective_function(lst_surgery, lst_surgeon, lst_room, lst_date):
    objective = pulp.lpSum(((d - A[surgery.get_surgery_id()]) + max(d - Due[surgery.get_surgery_id()], 0)) * x[surgery.get_surgery_id(), r, d] * U[surgery.get_surgery_id()] for surgery in lst_surgery for r in lst_room for d in lst_date)
    objective += pulp.lpSum(((Due[surgery.get_surgery_id()] - A[surgery.get_surgery_id()]) + (len(lst_date) + 1 - Due[surgery.get_surgery_id()])) * U[surgery.get_surgery_id()] for surgery in lst_surgery) - pulp.lpSum(((Due[surgery.get_surgery_id()] - A[surgery.get_surgery_id()]) + (len(lst_date) + 1 - Due[surgery.get_surgery_id()])) * U[surgery.get_surgery_id()] * x[surgery.get_surgery_id(), r, d] for surgery in lst_surgery for r in lst_room for d in lst_date)
    objective += 1000 * (pulp.lpSum(U[surgery.get_surgery_id()] for surgery in lst_surgery) - pulp.lpSum(U[surgery.get_surgery_id()] * x[surgery.get_surgery_id(), r, d] for surgery in lst_surgery for r in R for d in lst_date))
    objective += pulp.lpSum(n[surgeon.get_surgeon_id(), d] for surgeon in lst_surgeon for d in lst_date) + pulp.lpSum(ot[r, d] for r in lst_room for d in lst_date)
    objective += pulp.lpSum(wt[surgeon.get_surgeon_id(), d] for surgeon in lst_surgeon for d in lst_date) + pulp.lpSum(U[surgery.get_surgery_id()] * ts[surgery.get_surgery_id()] for surgery in lst_surgery)
    objective += pulp.lpSum(1 / dict_room_group_prob[g, r] * x[surgery.get_surgery_id(), r, d] for g in G for surgery in list(set(list_surgery) & set(dict_surgery_group[g])) for r in dict_room_group[g])
    return objective

print("目的関数値")
problem += objective_function(list_surgery, list_surgeon, R, D)

print("制約1")
for surgery in list_surgery:
    problem += pulp.lpSum(x[surgery.get_surgery_id(), r, d] for r in R for d in D) <= 1

print("制約2")
for surgery in list_surgery:
    problem += pulp.lpSum(q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d] for surgeon in list_surgeon for d in D) <= 1

print("制約3")
for surgery in list_surgery:
    for d in D:
        problem += pulp.lpSum(q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d] for surgeon in list_surgeon) == pulp.lpSum(x[surgery.get_surgery_id(), r, d] for r in R)

print("制約4")
for g in G:
    for surgery in dict_surgery_group[g]:
        for r in list(set(R) - set(dict_room_group[g])):
            for d in D:
                problem += x[surgery.get_surgery_id(), r, d] == 0

print("制約5")
for g in G:
    for surgery in dict_surgery_group[g]:
        for surgeon in list(set(list_surgeon) - set(dict_surgeon_group[g])):
            for d in D:
                problem += q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d] == 0

print("制約6")
for surgery in list_surgery:
    for d in D:
        problem += A[surgery.get_surgery_id()] * pulp.lpSum(x[surgery.get_surgery_id(), r, d] for r in R) <= d

print("制約7")
for surgeon in list_surgeon:
    for d in D:
        problem += pulp.lpSum(q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d] for surgery in list_surgery) <= len(list_surgery) * n[surgeon.get_surgeon_id(), d]

print("制約8")
for surgeon in list_surgeon:
    for d in D:
        problem += n[surgeon.get_surgeon_id(), d] <= pulp.lpSum(q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d] for surgery in list_surgery)

print("制約9")
for surgeon in list_surgeon:
    for d in D:
        problem += pulp.lpSum(q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d] for surgery in list_surgery) <= 3

print("制約10")
for surgery1 in list_surgery:
    for surgery2 in list_surgery:
        if surgery1 != surgery2:
            for r in R:
                for d in D:
                    problem += 2 * (y[surgery1.get_surgery_id(), surgery2.get_surgery_id(), r, d] + y[surgery2.get_surgery_id(), surgery1.get_surgery_id(), r, d]) <= x[surgery1.get_surgery_id(), r, d] + x[surgery2.get_surgery_id(), r, d]

print("制約11")
for surgery1 in list_surgery:
    for surgery2 in list_surgery:
        if surgery1 != surgery2:
            for r in R:
                for d in D:
                    problem += y[surgery1.get_surgery_id(), surgery2.get_surgery_id(), r, d] + y[surgery2.get_surgery_id(), surgery1.get_surgery_id(), r, d] >= x[surgery1.get_surgery_id(), r, d] + x[surgery2.get_surgery_id(), r, d] - 1

print("制約12")
for surgery1 in list_surgery:
    for surgery2 in list_surgery:
        if surgery1 != surgery2:
            for surgeon in list_surgeon:
                for d in D:
                    problem += 2 * (z[surgery1.get_surgery_id(), surgery2.get_surgery_id(), surgeon.get_surgeon_id(), d] + z[surgery2.get_surgery_id(), surgery1.get_surgery_id(), surgeon.get_surgeon_id(), d]) <= q[surgery1.get_surgery_id(), surgeon.get_surgeon_id(), d] + q[surgery2.get_surgery_id(), surgeon.get_surgeon_id(), d]

print("制約13")
for surgery1 in list_surgery:
    for surgery2 in list_surgery:
        if surgery1 != surgery2:
            for surgeon in list_surgeon:
                for d in D:
                    problem += z[surgery1.get_surgery_id(), surgery2.get_surgery_id(), surgeon.get_surgeon_id(), d] + z[surgery2.get_surgery_id(), surgery1.get_surgery_id(), surgeon.get_surgeon_id(), d] >= q[surgery1.get_surgery_id(), surgeon.get_surgeon_id(), d] + q[surgery2.get_surgery_id(), surgeon.get_surgeon_id(), d] - 1

print("制約14")
for surgery in list_surgery:
    problem += ts[surgery.get_surgery_id()] >= TP[surgery.get_surgery_id()]

print("制約15")
for surgery1 in list_surgery:
    for surgery2 in list_surgery:
        if surgery1 != surgery2:
            for r in R:
                for d in D:
                    problem += ts[surgery2.get_surgery_id()] - TP[surgery2.get_surgery_id()] >= ts[surgery1.get_surgery_id()] + TS[surgery1.get_surgery_id()] + TC[surgery1.get_surgery_id()] - M * (1 - y[surgery1.get_surgery_id(), surgery2.get_surgery_id(), r, d])

print("制約16")
for surgery1 in list_surgery:
    for surgery2 in list_surgery:
        if surgery1 != surgery2:
            for surgeon in list_surgeon:
                for d in D:
                    problem += ts[surgery2.get_surgery_id()] >= ts[surgery1.get_surgery_id()] + TS[surgery1.get_surgery_id()] + PT - M * (1 - z[surgery1.get_surgery_id(), surgery2.get_surgery_id(), surgeon.get_surgeon_id(), d])

print("制約17")
for surgery in list_surgery:
    for surgeon in list_surgeon:
        for d in D:
            problem += tsS[surgeon.get_surgeon_id(), d] <= ts[surgery.get_surgery_id()] + M * (1 - q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d])

print("制約18")
for surgery in list_surgery:
    for r in R:
        for d in D:
            problem += msR[r, d] >= ts[surgery.get_surgery_id()] + TS[surgery.get_surgery_id()] + TC[surgery.get_surgery_id()] - M * (1 - x[surgery.get_surgery_id(), r, d])

print("制約19")
for r in R:
    for d in D:
        problem += ot[r, d] >= msR[r, d] - T

print("制約20")
for surgery in list_surgery:
    for surgeon in list_surgeon:
        for d in D:
            problem += msS[surgeon.get_surgeon_id(), d] >= ts[surgery.get_surgery_id()] + TS[surgery.get_surgery_id()] - M * (1 - q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d])

print("制約21")
for surgeon in list_surgeon:
    for d in D:
        problem += wt[surgeon.get_surgeon_id(), d] >= msS[surgeon.get_surgeon_id(), d] - tsS[surgeon.get_surgeon_id(), d] - pulp.lpSum(TS[surgery.get_surgery_id()] * q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d] for surgery in list_surgery)

print("制約22")
for r in R:
    for d in D:
        problem += ot[r, d] <= 120


print("式")
print("--------")
print(problem)
print("--------")
print("")
solver = pulp.solvers.CPLEX_CMD(timelimit=100)
print("計算開始")
time_start = time.time()
result_status = problem.solve(solver)
time_stop = time.time()
print("計算結果")
print("********")

print("最適性 = {:}, 目的関数値 = {:}, 計算時間 = {:} (秒)"
      .format(pulp.LpStatus[result_status], pulp.value(problem.objective),
              time_stop - time_start))

overtime = sum(ot[r, d].value() for d in D for r in R)
print("残業時間 = {:}".format(overtime))
print("手術数={:}".format(len(list_surgery)))
print("外科医の人数={:}".format(len(list_surgeon)))
print("手術室の数={:}".format(len(R)))
print("計画日数={:}".format(len(D)))
def sort(lst):
    for i in range(len(lst) - 1):
        for j in range(i + 1, len(lst)):
            if ts[lst[i].get_surgery_id()].value() > ts[lst[j].get_surgery_id()].value():
                tmp = lst[i]
                lst[i] = lst[j]
                lst[j] = tmp
    return lst

def classify_room_date(lst, r, d):
    lst_result = []
    for surgery in lst:
        if x[surgery.get_surgery_id(), r, d].value() == 1:
            lst_result.append(surgery)
    return sort(lst_result)

def classify_surgeon_date(lst, surgeon, d):
    lst_result = []
    for surgery in lst:
        if q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d].value() == 1:
            lst_result.append(surgery)
    return sort(lst_result)


def get_total_time(surgery):
    return TP[surgery.get_surgery_id()] + TS[surgery.get_surgery_id()] + TC[surgery.get_surgery_id()]

def get_entry_time(surgery):
    return ts[surgery.get_surgery_id()].value() - TP[surgery.get_surgery_id()]

def get_end_time(surgery):
    return ts[surgery.get_surgery_id()].value() + TS[surgery.get_surgery_id()]

def get_exit_time(surgery):
    return get_end_time(surgery) + TC[surgery.get_surgery_id()]
def create_excel_file(sheet, list_surgery, list_room, d):
    length = 0
    for i in range(len(list_room)):
        for j in range(len(list_surgery) * 4):
            sheet.cell(row=j + 2, column=int(list_room[i]) + 1).value = 0
    for r in list_room:
        list_surgery_in_r = classify_room_date(list_surgery, r, d)
        sheet.cell(row=1, column=int(r) + 1).value = "手術室" + r
        for i in range(len(list_surgery_in_r) * 4):
            surgery = list_surgery_in_r[int(i / 4)]
            if i % 4 == 0:
                sheet.cell(row=i + 2 + length, column=1).value = "空白"
                if i == 0:
                    sheet.cell(row=i + 2 + length, column=int(r) + 1).value = ts[surgery.get_surgery_id()].value() - TP[surgery.get_surgery_id()]
                else:
                    prev_surgery = list_surgery_in_r[int(i / 4 - 1)]
                    sheet.cell(row=i + 2 + length, column=int(r) + 1).value = get_entry_time(surgery) - get_exit_time(prev_surgery)
            elif i % 4 == 1:
                sheet.cell(row=i + 2 + length, column=1).value = "準備" + str(surgery.get_surgery_id())
                sheet.cell(row=i + 2 + length, column=int(r) + 1).value = TP[surgery.get_surgery_id()]
            elif i % 4 == 2:
                sheet.cell(row=i + 2 + length, column=1).value = "手術" + str(surgery.get_surgery_id())
                sheet.cell(row=i + 2 + length, column=int(r) + 1).value = TS[surgery.get_surgery_id()]
            if i % 4 == 3:
                sheet.cell(row=i + 2 + length, column=1).value = "清掃" + str(surgery.get_surgery_id())
                sheet.cell(row=i + 2 + length, column=int(r) + 1).value = TC[surgery.get_surgery_id()]
        length += len(list_surgery_in_r) * 4

def create_excel_file2(sheet, list_surgery, list_surgeon, d):
    length = 0
    for i in range(len(list_surgeon)):
        for j in range(len(list_surgery) * 2):
            sheet.cell(row=j + 2, column=i + 2).value = 0
    for i in range(len(list_surgeon)):
        surgeon = list_surgeon[i]
        list_surgery_by_surgeon = classify_surgeon_date(list_surgery, surgeon, d)
        sheet.cell(row=1, column=i + 2).value = "外科医" + str(surgeon.get_surgeon_id())
        for j in range(len(list_surgery_by_surgeon) * 2):
            surgery = list_surgery_by_surgeon[int(j / 2)]
            if j % 2 == 0:
                sheet.cell(row=j + 2 + length, column=1).value = "空白"
                if j == 0:
                    sheet.cell(row=j + 2 + length, column=i + 2).value = ts[surgery.get_surgery_id()].value()
                else:
                    prev_surgery = list_surgery_by_surgeon[int(j / 2 - 1)]
                    sheet.cell(row=j + 2 + length, column=i + 2).value = ts[surgery.get_surgery_id()].value() - get_end_time(prev_surgery)
            else:
                sheet.cell(row=j + 2 + length, column=1).value = "手術" + surgery.get_surgery_id()
                sheet.cell(row=j + 2 + length, column=i + 2).value = TS[surgery.get_surgery_id()]
        length += len(list_surgery_by_surgeon) * 2

book = openpyxl.Workbook()
book2 = openpyxl.Workbook()

for d in D:
    scheduled_list = []
    scheduled_surgeon_list = []
    for surgery in list_surgery:
        for r in R:
            if x[surgery.get_surgery_id(), r, d].value() == 1:
                scheduled_list.append(surgery)
        for surgeon in list_surgeon:
            if q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d].value() == 1:
                scheduled_surgeon_list.append(surgeon)
    scheduled_surgeon_list = list(set(scheduled_surgeon_list))
    sheet = book.worksheets[d - 1]
    sheet2 = book2.worksheets[d - 1]
    sheet.title = 'surgery_schedule_orps' + str(d)
    sheet2.title = 'surgeon_schedule_orps' + str(d)
    if D.index(d) < len(D) - 1:
        book.create_sheet()
        book2.create_sheet()
    create_excel_file(sheet, scheduled_list, R, d)
    create_excel_file2(sheet2, scheduled_list, scheduled_surgeon_list, d)
book.save(url)
book.close()
book2.save(url2)
book2.close()
graph.create_ganttchart(url)
graph.create_surgeon_gantt_chart(url2)
def to_time(num):
    hour = int(num / 60)
    minute = int(num % 60)
    if hour < 10:
        hour = "0" + str(hour)
    if minute < 10:
        minute = "0" + str(minute)
    return str(hour) + ":" + str(minute)

def to_time_of_day(num):
    hour = int(num / 60) + 9
    minute = int(num % 60)
    if hour < 10:
        hour = "0" + str(hour)
    if minute < 10:
        minute = "0" + str(minute)
    return str(hour) + ":" + str(minute)
def display(lst_surgery, lst_room, lst_surgeon, lst_date):
    for surgery in lst_surgery:
        for r in lst_room:
            for surgeon in lst_surgeon:
                for d in lst_date:
                    start_int = ts[surgery.get_surgery_id()].value()
                    entry_int = start_int - TP[surgery.get_surgery_id()]
                    end_int = start_int + TS[surgery.get_surgery_id()]
                    exit_int = end_int + TC[surgery.get_surgery_id()]
                    entry_time = to_time_of_day(entry_int)
                    start_time = to_time_of_day(start_int)
                    end_time = to_time_of_day(end_int)
                    exit_time = to_time_of_day(exit_int)
                    if x[surgery.get_surgery_id(), r, d].value() == 1 and q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d].value() == 1:
                        string = str(surgery.get_surgery_id()).ljust(6) + "\t" + str(surgeon.get_surgeon_id()).ljust(6) + "\t" + str(r).ljust(6) + "\t" + surgery.get_group().ljust(6) + "\t" + str(d).ljust(6)
                        string += "\t" + to_time(TP[surgery.get_surgery_id()]).ljust(6) + "\t" + to_time(TS[surgery.get_surgery_id()]).ljust(6) + "\t" + to_time(TC[surgery.get_surgery_id()]).ljust(6)
                        string += "\t" + entry_time.ljust(6) + "\t" + start_time.ljust(6) + "\t" + end_time.ljust(6) + "\t" + exit_time.ljust(6)
                        print(string)

def print_subject():
    string = "surgery".ljust(6) + "\t" + "surgeon".ljust(6) + "\t" + "room".ljust(6) + "\t" + "group".ljust(12) + "\t" + "date"
    string += "\t" + "pre".ljust(6) + "\t" + "sur".ljust(6) + "\t" + "cle".ljust(6)
    string += "\t" + "entry".ljust(6) + "\t" + "start".ljust(6) + "\t" + "end".ljust(6) + "\t" + "exit".ljust(6)
    print(string)
"""
for d in D:
    for r in R:
        list_surgery_room = classfy_room_date(list_surgery, r, d)
        if len(list_surgery_room) > 0:
            print("date = {:}".format(d))
            print("手術室" + str(r))
            print_subject()
            display(list_surgery_room, R, list_surgeon, D)
            print()
"""
"""
for d in D:
    for surgeon in list_surgeon:
        list_surgery_surgeon = classify_surgeon_date(list_surgery, surgeon, d)
        if len(list_surgery_surgeon) > 0:
            print("date = {:}".format(d))
            print("外科医" + str(surgeon.get_surgeon_id()))
            print_subject()
            display(list_surgery_surgeon, R, list_surgeon, D)
            print()"""


with open('solution.csv', 'w') as csv_file:
    fieldnames = ['surgery id', 'surgeon id', 'room', 'group', 'date', 'release date', 'due date', 'preparation time', 'surgery time', 'cleaning time', 'entry time', 'start time', 'end time', 'exit time']
    writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
    writer.writeheader()
    for d in D:
        for r in R:
            list_surgery_room = classify_room_date(list_surgery, r, d)
            for surgery in list_surgery_room:
                for surgeon in list_surgeon:
                    if q[surgery.get_surgery_id(), surgeon.get_surgeon_id(), d].value() == 1 and x[surgery.get_surgery_id(), r, d].value() == 1:
                        start_int = ts[surgery.get_surgery_id()].value()
                        entry_int = start_int - TP[surgery.get_surgery_id()]
                        end_int = start_int + TS[surgery.get_surgery_id()]
                        exit_int = end_int + TC[surgery.get_surgery_id()]
                        entry_time = to_time_of_day(entry_int)
                        start_time = to_time_of_day(start_int)
                        end_time = to_time_of_day(end_int)
                        exit_time = to_time_of_day(exit_int)
                        writer.writerow({'surgery id': surgery.get_surgery_id(), 'surgeon id': surgeon.get_surgeon_id(), 'room': r, 'group': surgery.get_group(), 'date': d, 'release date': surgery.get_release_date(), 'due date': surgery.get_due_date(), 'preparation time': TP[surgery.get_surgery_id()], 'surgery time': TS[surgery.get_surgery_id()], 'cleaning time': TC[surgery.get_surgery_id()], 'entry time': entry_time, 'start time': start_time, 'end time': end_time, 'exit time': exit_time})

"""for surgery in list_surgery:
    print("ts[{:}]={:}".format(surgery.get_surgery_id(), ts[surgery1.get_surgery_id()].value()))
"""
